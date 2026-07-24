
import base64
import hashlib
import hmac
import secrets
import re
import sqlite3
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACIÓN
# ============================================================

st.set_page_config(
    page_title="Jet Airways Academy | Sistema de Ventas",
    page_icon="✈️",
    layout="wide",
)


BACKGROUND_PATH = (
    Path(__file__).resolve().parent / "assets" / "fondo_avion.png"
)


def cargar_fondo_css():
    if not BACKGROUND_PATH.exists():
        return ""
    contenido = base64.b64encode(BACKGROUND_PATH.read_bytes()).decode("utf-8")
    return f"data:image/png;base64,{contenido}"


# ============================================================
# IDENTIDAD VISUAL JET AIRWAYS ACADEMY
# ============================================================

def aplicar_estilo_corporativo():
    fondo_css = cargar_fondo_css()
    estilos = """
        <style>
        :root {
            --jet-navy: #04152F;
            --jet-navy-2: #08264D;
            --jet-blue: #0E4A92;
            --jet-blue-bright: #2A6EDB;
            --jet-red: #E30613;
            --jet-red-hover: #B9040E;
            --jet-gold: #D8B46A;
            --jet-white: #FFFFFF;
            --jet-muted: #AFC3DF;
            --jet-panel: rgba(9, 35, 74, 0.88);
            --jet-panel-soft: rgba(255, 255, 255, 0.07);
        }

        html, body, [class*="css"] {
            font-family: "Inter", "Segoe UI", Arial, sans-serif;
        }

        .stApp {
            background:
                linear-gradient(
                    rgba(2, 10, 24, 0.88),
                    rgba(3, 19, 41, 0.91)
                ),
                url("__FONDO_AVION__") center center / cover fixed no-repeat;
            color: var(--jet-white);
        }

        [data-testid="stHeader"] {
            background: rgba(3, 13, 31, 0.76);
            backdrop-filter: blur(12px);
            border-bottom: 1px solid rgba(255,255,255,0.06);
        }

        [data-testid="stToolbar"] {
            right: 1rem;
        }

        [data-testid="stSidebar"] {
            background:
                linear-gradient(180deg, rgba(3, 13, 31, 0.98) 0%, rgba(5, 28, 61, 0.98) 100%);
            border-right: 1px solid rgba(255,255,255,0.09);
            box-shadow: 18px 0 45px rgba(0,0,0,0.22);
        }

        [data-testid="stSidebar"] * {
            color: var(--jet-white);
        }

        [data-testid="stSidebar"] [role="radiogroup"] label {
            background: rgba(255,255,255,0.045);
            border: 1px solid rgba(255,255,255,0.06);
            border-radius: 12px;
            padding: 0.62rem 0.72rem;
            margin-bottom: 0.38rem;
            transition: all 0.18s ease;
        }

        [data-testid="stSidebar"] [role="radiogroup"] label:hover {
            background: rgba(42,110,219,0.18);
            border-color: rgba(91,145,228,0.42);
            transform: translateX(2px);
        }

        h1, h2, h3, h4, h5, h6, p, label,
        [data-testid="stCaptionContainer"] {
            color: var(--jet-white) !important;
        }

        h1 {
            letter-spacing: -0.035em;
            font-weight: 850 !important;
        }

        h2, h3 {
            letter-spacing: -0.02em;
        }

        .block-container {
            max-width: 1450px;
            padding-top: 1.55rem;
            padding-bottom: 2.5rem;
        }

        div[data-testid="stForm"] {
            background:
                linear-gradient(145deg, rgba(11, 42, 88, 0.94), rgba(5, 24, 52, 0.94));
            border: 1px solid rgba(255, 255, 255, 0.12);
            border-radius: 22px;
            padding: 1.55rem 1.65rem 1.25rem 1.65rem;
            box-shadow:
                0 28px 70px rgba(0, 0, 0, 0.34),
                inset 0 1px 0 rgba(255,255,255,0.05);
        }

        div[data-testid="stTextInput"] input,
        div[data-testid="stNumberInput"] input,
        div[data-testid="stDateInput"] input,
        div[data-testid="stTimeInput"] input,
        div[data-baseweb="select"] > div,
        div[data-testid="stFileUploader"] section,
        div[data-testid="stCameraInput"] > div {
            background: rgba(255, 255, 255, 0.97) !important;
            color: #071A3E !important;
            border-radius: 12px !important;
            border: 1px solid rgba(7,26,62,0.10) !important;
            box-shadow: 0 6px 20px rgba(0,0,0,0.06);
        }

        div[data-testid="stTextInput"] input:focus,
        div[data-testid="stNumberInput"] input:focus,
        div[data-testid="stDateInput"] input:focus {
            border-color: #2A6EDB !important;
            box-shadow: 0 0 0 3px rgba(42,110,219,0.14) !important;
        }

        div[data-testid="stTextInput"] input::placeholder {
            color: #778399 !important;
        }

        .stButton > button,
        .stDownloadButton > button,
        div[data-testid="stFormSubmitButton"] button {
            background:
                linear-gradient(90deg, #C90714 0%, #E30613 48%, #F03B45 100%) !important;
            color: white !important;
            border: none !important;
            border-radius: 12px !important;
            font-weight: 800 !important;
            min-height: 3rem;
            letter-spacing: 0.01em;
            box-shadow: 0 12px 30px rgba(227, 6, 19, 0.26);
            transition: all 0.18s ease;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover,
        div[data-testid="stFormSubmitButton"] button:hover {
            background: linear-gradient(90deg, #A9040D 0%, #D10915 100%) !important;
            transform: translateY(-2px);
            box-shadow: 0 16px 34px rgba(227, 6, 19, 0.32);
        }

        div[data-testid="stMetric"] {
            background:
                linear-gradient(145deg, rgba(255,255,255,0.10), rgba(255,255,255,0.045));
            border: 1px solid rgba(255, 255, 255, 0.11);
            border-radius: 18px;
            padding: 1.05rem 1.1rem;
            box-shadow: 0 18px 45px rgba(0,0,0,0.18);
            min-height: 112px;
        }

        div[data-testid="stMetric"] label {
            color: #BFD0E8 !important;
            font-size: 0.82rem !important;
            font-weight: 700 !important;
            text-transform: uppercase;
            letter-spacing: 0.055em;
        }

        div[data-testid="stMetricValue"] {
            color: white !important;
            font-weight: 850 !important;
            letter-spacing: -0.025em;
        }

        div[data-testid="stDataFrame"] {
            background: white;
            border-radius: 16px;
            overflow: hidden;
            box-shadow: 0 20px 50px rgba(0,0,0,0.20);
            border: 1px solid rgba(255,255,255,0.12);
        }

        [data-testid="stExpander"] {
            background: linear-gradient(145deg, rgba(9, 36, 80, 0.86), rgba(5,24,52,0.86));
            border: 1px solid rgba(255, 255, 255, 0.11);
            border-radius: 15px;
            box-shadow: 0 12px 30px rgba(0,0,0,0.14);
        }

        div[data-testid="stAlert"] {
            border-radius: 14px;
        }

        .jet-brand-title {
            text-align: center;
            color: white;
            font-size: 2.25rem;
            font-weight: 900;
            margin-top: 0.2rem;
            margin-bottom: 0.2rem;
            letter-spacing: -0.035em;
        }

        .jet-brand-subtitle {
            text-align: center;
            color: #B9CBE4;
            font-size: 1rem;
            margin-bottom: 1.5rem;
        }

        .jet-footer {
            text-align: center;
            color: rgba(255,255,255,0.55);
            font-size: 0.78rem;
            margin-top: 2rem;
        }

        .jet-hero {
            position: relative;
            overflow: hidden;
            border-radius: 24px;
            padding: 1.65rem 1.9rem;
            margin-bottom: 1.4rem;
            background:
                linear-gradient(120deg, rgba(6,27,66,0.96) 0%, rgba(11,56,115,0.90) 62%, rgba(227,6,19,0.80) 140%);
            border: 1px solid rgba(255,255,255,0.13);
            box-shadow: 0 28px 70px rgba(0,0,0,0.31);
        }

        .jet-hero::after {
            content: "✈";
            position: absolute;
            right: 2.4rem;
            top: 0.35rem;
            font-size: 7.8rem;
            color: rgba(255,255,255,0.075);
            transform: rotate(-8deg);
        }

        .jet-kicker {
            color: #D6E3F6;
            text-transform: uppercase;
            letter-spacing: 0.16em;
            font-size: 0.72rem;
            font-weight: 800;
            margin-bottom: 0.5rem;
        }

        .jet-hero-title {
            color: white;
            font-size: 2rem;
            font-weight: 900;
            letter-spacing: -0.035em;
            line-height: 1.05;
            margin-bottom: 0.55rem;
        }

        .jet-hero-copy {
            color: #C6D7EE;
            font-size: 0.98rem;
            max-width: 760px;
            margin: 0;
        }

        .jet-badge {
            display: inline-block;
            margin-top: 0.95rem;
            padding: 0.42rem 0.75rem;
            border-radius: 999px;
            background: rgba(255,255,255,0.10);
            border: 1px solid rgba(255,255,255,0.16);
            color: #F3F7FD;
            font-size: 0.78rem;
            font-weight: 750;
        }

        .jet-section-title {
            display: flex;
            align-items: center;
            gap: 0.7rem;
            margin: 0.2rem 0 1rem 0;
            font-size: 1.2rem;
            font-weight: 850;
            color: white;
        }

        .jet-section-title::before {
            content: "";
            width: 4px;
            height: 24px;
            border-radius: 10px;
            background: linear-gradient(180deg, #E30613, #FF6670);
            box-shadow: 0 0 16px rgba(227,6,19,0.45);
        }

        .status-chip {
            display: inline-block;
            padding: 0.32rem 0.62rem;
            border-radius: 999px;
            font-size: 0.76rem;
            font-weight: 800;
            border: 1px solid transparent;
        }

        .status-pendiente {
            background: rgba(245, 158, 11, 0.16);
            color: #FFD98A;
            border-color: rgba(245, 158, 11, 0.30);
        }

        .status-verificando {
            background: rgba(59, 130, 246, 0.16);
            color: #A9CCFF;
            border-color: rgba(59, 130, 246, 0.30);
        }

        .status-aprobado {
            background: rgba(34, 197, 94, 0.16);
            color: #A8F0BF;
            border-color: rgba(34, 197, 94, 0.30);
        }

        .status-rechazado {
            background: rgba(239, 68, 68, 0.16);
            color: #FFB4B4;
            border-color: rgba(239, 68, 68, 0.30);
        }

        .sidebar-brand {
            text-align: center;
            padding: 0.45rem 0 1rem 0;
        }

        .sidebar-brand-icon {
            width: 54px;
            height: 54px;
            border-radius: 16px;
            margin: 0 auto 0.7rem auto;
            display: flex;
            align-items: center;
            justify-content: center;
            background: linear-gradient(145deg, #164A91, #0A2852);
            border: 1px solid rgba(255,255,255,0.12);
            box-shadow: 0 14px 30px rgba(0,0,0,0.24);
            font-size: 1.7rem;
        }

        .sidebar-brand-name {
            font-size: 1rem;
            font-weight: 850;
            letter-spacing: -0.02em;
        }

        .sidebar-brand-sub {
            font-size: 0.72rem;
            color: #AFC3DF;
            margin-top: 0.18rem;
        }

        @media (max-width: 900px) {
            .jet-hero-title {
                font-size: 1.55rem;
            }

            .jet-hero::after {
                font-size: 5.5rem;
                right: 0.7rem;
            }

            .block-container {
                padding-left: 0.75rem;
                padding-right: 0.75rem;
            }
        }
        </style>
        """
    estilos = estilos.replace("__FONDO_AVION__", fondo_css)
    st.markdown(estilos, unsafe_allow_html=True)


aplicar_estilo_corporativo()

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "ventas.db"
LOGO_PATH = BASE_DIR / "assets" / "logo_jet_airways.png"

TIPO_CAMBIO_INICIAL = 6.96

CURSOS_INICIALES = {
    "PPL": {"nombre": "PPL", "monto": 100.0, "moneda": "USD"},
    "PPC": {"nombre": "PPC", "monto": 275.0, "moneda": "USD"},
    "IFR": {"nombre": "IFR", "monto": 275.0, "moneda": "USD"},
    "TCP": {"nombre": "TCP", "monto": 50.0, "moneda": "USD"},
    "EOU": {"nombre": "EOU", "monto": 30.0, "moneda": "USD"},
    "AGT": {"nombre": "AGT", "monto": 120.0, "moneda": "BOB"},
    "AGR": {"nombre": "AGR", "monto": 140.0, "moneda": "BOB"},
    "MTC": {"nombre": "MTC", "monto": 60.0, "moneda": "USD"},
}

SEDES_INICIALES = ["La Paz", "El Alto", "Santa Cruz", "Cochabamba"]

BANCOS_INICIALES = [
    "Banco Nacional de Bolivia (BNB)",
    "Banco Mercantil Santa Cruz",
    "Banco de Crédito BCP",
    "Banco Bisa",
    "Banco Unión",
    "Banco Económico",
    "Banco Ganadero",
    "QR",
    "Efectivo",
    "Otro",
]

ROLES = ["Administrador", "Supervisor", "Vendedor"]


# ============================================================
# BASE DE DATOS
# ============================================================

def conectar():
    return sqlite3.connect(DB_PATH, check_same_thread=False)


def crear_base_datos():
    """Crea o actualiza las tablas sin borrar las ventas existentes."""
    with conectar() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS ventas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                curso TEXT NOT NULL,
                monto REAL NOT NULL,
                cliente TEXT NOT NULL,
                vendedor TEXT NOT NULL,
                comision_original REAL NOT NULL,
                moneda_comision TEXT NOT NULL,
                tipo_cambio REAL NOT NULL,
                comision_bs REAL NOT NULL,
                ingreso_neto_bs REAL NOT NULL,
                creado_en TEXT NOT NULL
            )
            """
        )

        columnas_ventas = {
            fila[1] for fila in conn.execute("PRAGMA table_info(ventas)").fetchall()
        }
        if "creado_por" not in columnas_ventas:
            conn.execute(
                "ALTER TABLE ventas ADD COLUMN creado_por TEXT NOT NULL DEFAULT ''"
            )

        if "sede" not in columnas_ventas:
            conn.execute(
                "ALTER TABLE ventas ADD COLUMN sede TEXT NOT NULL DEFAULT ''"
            )

        # Campos de control del pago y comprobante.
        columnas_nuevas = {
            "numero_transaccion": "TEXT",
            "banco": "TEXT",
            "fecha_hora_pago": "TEXT",
            "estado_pago": "TEXT NOT NULL DEFAULT 'Pendiente'",
            "comprobante": "BLOB",
            "comprobante_mime": "TEXT",
            "comprobante_nombre": "TEXT",
        }

        for nombre_columna, tipo_columna in columnas_nuevas.items():
            if nombre_columna not in columnas_ventas:
                conn.execute(
                    f"ALTER TABLE ventas ADD COLUMN {nombre_columna} {tipo_columna}"
                )

        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_ventas_numero_transaccion
            ON ventas(numero_transaccion)
            WHERE numero_transaccion IS NOT NULL
              AND TRIM(numero_transaccion) <> ''
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                usuario TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                salt TEXT NOT NULL,
                rol TEXT NOT NULL,
                activo INTEGER NOT NULL DEFAULT 1,
                creado_en TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS cursos (
                codigo TEXT PRIMARY KEY,
                nombre TEXT NOT NULL,
                comision REAL NOT NULL,
                moneda TEXT NOT NULL,
                activo INTEGER NOT NULL DEFAULT 1,
                creado_en TEXT NOT NULL,
                actualizado_en TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sedes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE,
                activo INTEGER NOT NULL DEFAULT 1,
                creado_en TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bancos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL UNIQUE,
                activo INTEGER NOT NULL DEFAULT 1,
                creado_en TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS configuracion (
                clave TEXT PRIMARY KEY,
                valor TEXT NOT NULL,
                actualizado_en TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS alumnos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL UNIQUE,
                nombre_completo TEXT NOT NULL,
                documento TEXT NOT NULL,
                telefono TEXT NOT NULL,
                correo TEXT,
                numero_transaccion TEXT NOT NULL,
                venta_id INTEGER,
                curso TEXT,
                sede TEXT,
                monto_reportado REAL,
                estado TEXT NOT NULL DEFAULT 'Pendiente',
                observacion TEXT,
                registrado_en TEXT NOT NULL,
                revisado_por TEXT,
                revisado_en TEXT,
                FOREIGN KEY (venta_id) REFERENCES ventas(id)
            )
            """
        )

        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_alumnos_transaccion
            ON alumnos(numero_transaccion)
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS auditoria_datos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario TEXT NOT NULL,
                accion TEXT NOT NULL,
                detalle TEXT,
                creado_en TEXT NOT NULL
            )
            """
        )

        ahora = datetime.now().isoformat(timespec="seconds")

        cantidad_cursos = conn.execute("SELECT COUNT(*) FROM cursos").fetchone()[0]
        if cantidad_cursos == 0:
            conn.executemany(
                """
                INSERT INTO cursos (
                    codigo, nombre, comision, moneda, activo,
                    creado_en, actualizado_en
                )
                VALUES (?, ?, ?, ?, 1, ?, ?)
                """,
                [
                    (
                        codigo,
                        datos["nombre"],
                        float(datos["monto"]),
                        datos["moneda"],
                        ahora,
                        ahora,
                    )
                    for codigo, datos in CURSOS_INICIALES.items()
                ],
            )

        cantidad_sedes = conn.execute("SELECT COUNT(*) FROM sedes").fetchone()[0]
        if cantidad_sedes == 0:
            conn.executemany(
                "INSERT INTO sedes (nombre, activo, creado_en) VALUES (?, 1, ?)",
                [(nombre, ahora) for nombre in SEDES_INICIALES],
            )

        cantidad_bancos = conn.execute("SELECT COUNT(*) FROM bancos").fetchone()[0]
        if cantidad_bancos == 0:
            conn.executemany(
                "INSERT INTO bancos (nombre, activo, creado_en) VALUES (?, 1, ?)",
                [(nombre, ahora) for nombre in BANCOS_INICIALES],
            )

        conn.execute(
            """
            INSERT OR IGNORE INTO configuracion (clave, valor, actualizado_en)
            VALUES ('tipo_cambio_usd_bs', ?, ?)
            """,
            (str(TIPO_CAMBIO_INICIAL), ahora),
        )
        conn.commit()



def cargar_cursos(solo_activos=False):
    consulta = """
        SELECT codigo, nombre, comision, moneda, activo,
               creado_en, actualizado_en
        FROM cursos
    """
    if solo_activos:
        consulta += " WHERE activo = 1"
    consulta += " ORDER BY nombre, codigo"

    with conectar() as conn:
        return pd.read_sql_query(consulta, conn)


def obtener_curso(codigo):
    with conectar() as conn:
        fila = conn.execute(
            """
            SELECT codigo, nombre, comision, moneda, activo
            FROM cursos
            WHERE codigo = ?
            """,
            (codigo,),
        ).fetchone()

    if fila is None:
        raise KeyError(f"El curso {codigo} no existe.")

    return {
        "codigo": fila[0],
        "nombre": fila[1],
        "monto": float(fila[2]),
        "moneda": fila[3],
        "activo": bool(fila[4]),
    }


class CursosDinamicos:
    """Compatibilidad con la lógica existente, leyendo cursos desde SQLite."""

    def __getitem__(self, codigo):
        datos = obtener_curso(codigo)
        return {"monto": datos["monto"], "moneda": datos["moneda"]}

    def keys(self):
        return cargar_cursos(solo_activos=True)["codigo"].tolist()

    def items(self):
        cursos = cargar_cursos(solo_activos=False)
        return [
            (
                fila.codigo,
                {"monto": float(fila.comision), "moneda": fila.moneda},
            )
            for fila in cursos.itertuples()
        ]

    def __iter__(self):
        return iter(self.keys())

    def __contains__(self, codigo):
        try:
            obtener_curso(codigo)
            return True
        except KeyError:
            return False


COMISIONES = CursosDinamicos()


def crear_curso(codigo, nombre, comision, moneda, activo=True):
    codigo = codigo.strip().upper()
    nombre = nombre.strip()

    if not codigo or not nombre:
        raise ValueError("El código y el nombre del curso son obligatorios.")
    if comision < 0:
        raise ValueError("La comisión no puede ser negativa.")
    if moneda not in ["USD", "BOB"]:
        raise ValueError("La moneda debe ser USD o BOB.")

    ahora = datetime.now().isoformat(timespec="seconds")
    try:
        with conectar() as conn:
            conn.execute(
                """
                INSERT INTO cursos (
                    codigo, nombre, comision, moneda, activo,
                    creado_en, actualizado_en
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    codigo,
                    nombre,
                    float(comision),
                    moneda,
                    1 if activo else 0,
                    ahora,
                    ahora,
                ),
            )
            conn.commit()
    except sqlite3.IntegrityError as error:
        raise ValueError("Ya existe un curso con ese código.") from error


def actualizar_curso(
    codigo_actual,
    nuevo_codigo,
    nombre,
    comision,
    moneda,
    activo,
):
    codigo_actual = codigo_actual.strip().upper()
    nuevo_codigo = nuevo_codigo.strip().upper()
    nombre = nombre.strip()

    if not nuevo_codigo or not nombre:
        raise ValueError("El código y el nombre son obligatorios.")
    if comision < 0:
        raise ValueError("La comisión no puede ser negativa.")
    if moneda not in ["USD", "BOB"]:
        raise ValueError("La moneda debe ser USD o BOB.")

    ahora = datetime.now().isoformat(timespec="seconds")

    try:
        with conectar() as conn:
            conn.execute("BEGIN")
            if nuevo_codigo != codigo_actual:
                existe = conn.execute(
                    "SELECT 1 FROM cursos WHERE codigo = ?",
                    (nuevo_codigo,),
                ).fetchone()
                if existe:
                    raise ValueError("Ya existe otro curso con ese código.")

                # Se actualiza también el historial para conservar reportes coherentes.
                conn.execute(
                    "UPDATE ventas SET curso = ? WHERE curso = ?",
                    (nuevo_codigo, codigo_actual),
                )

            conn.execute(
                """
                UPDATE cursos
                SET codigo = ?, nombre = ?, comision = ?, moneda = ?,
                    activo = ?, actualizado_en = ?
                WHERE codigo = ?
                """,
                (
                    nuevo_codigo,
                    nombre,
                    float(comision),
                    moneda,
                    1 if activo else 0,
                    ahora,
                    codigo_actual,
                ),
            )
            conn.commit()
    except sqlite3.IntegrityError as error:
        raise ValueError("No se pudo actualizar el curso.") from error


def cantidad_ventas_curso(codigo):
    with conectar() as conn:
        return int(
            conn.execute(
                "SELECT COUNT(*) FROM ventas WHERE curso = ?",
                (codigo,),
            ).fetchone()[0]
        )


def eliminar_curso(codigo):
    cantidad = cantidad_ventas_curso(codigo)
    if cantidad > 0:
        raise ValueError(
            f"Este curso tiene {cantidad} venta(s). "
            "Desactívalo en lugar de eliminarlo."
        )

    with conectar() as conn:
        conn.execute("DELETE FROM cursos WHERE codigo = ?", (codigo,))
        conn.commit()


def cargar_catalogo(tabla, solo_activos=False):
    if tabla not in {"sedes", "bancos"}:
        raise ValueError("Catálogo no permitido.")

    consulta = f"SELECT id, nombre, activo, creado_en FROM {tabla}"
    if solo_activos:
        consulta += " WHERE activo = 1"
    consulta += " ORDER BY nombre"

    with conectar() as conn:
        return pd.read_sql_query(consulta, conn)


def crear_elemento_catalogo(tabla, nombre):
    if tabla not in {"sedes", "bancos"}:
        raise ValueError("Catálogo no permitido.")

    nombre = nombre.strip()
    if not nombre:
        raise ValueError("El nombre es obligatorio.")

    try:
        with conectar() as conn:
            conn.execute(
                f"""
                INSERT INTO {tabla} (nombre, activo, creado_en)
                VALUES (?, 1, ?)
                """,
                (nombre, datetime.now().isoformat(timespec="seconds")),
            )
            conn.commit()
    except sqlite3.IntegrityError as error:
        raise ValueError("Ese nombre ya existe.") from error


def actualizar_elemento_catalogo(tabla, id_elemento, nombre, activo):
    if tabla not in {"sedes", "bancos"}:
        raise ValueError("Catálogo no permitido.")

    nombre = nombre.strip()
    if not nombre:
        raise ValueError("El nombre es obligatorio.")

    try:
        with conectar() as conn:
            conn.execute(
                f"UPDATE {tabla} SET nombre = ?, activo = ? WHERE id = ?",
                (nombre, 1 if activo else 0, int(id_elemento)),
            )
            conn.commit()
    except sqlite3.IntegrityError as error:
        raise ValueError("Ese nombre ya existe.") from error


def eliminar_elemento_catalogo(tabla, id_elemento):
    if tabla not in {"sedes", "bancos"}:
        raise ValueError("Catálogo no permitido.")

    with conectar() as conn:
        fila = conn.execute(
            f"SELECT nombre FROM {tabla} WHERE id = ?",
            (int(id_elemento),),
        ).fetchone()

        if fila is None:
            raise ValueError("El elemento seleccionado ya no existe.")

        nombre = fila[0]
        columna_venta = "sede" if tabla == "sedes" else "banco"
        cantidad = conn.execute(
            f"SELECT COUNT(*) FROM ventas WHERE {columna_venta} = ?",
            (nombre,),
        ).fetchone()[0]

        if cantidad > 0:
            raise ValueError(
                f"Este elemento aparece en {cantidad} venta(s). "
                "Desactívalo en lugar de eliminarlo."
            )

        conn.execute(f"DELETE FROM {tabla} WHERE id = ?", (int(id_elemento),))
        conn.commit()


def obtener_tipo_cambio():
    with conectar() as conn:
        fila = conn.execute(
            """
            SELECT valor FROM configuracion
            WHERE clave = 'tipo_cambio_usd_bs'
            """
        ).fetchone()

    return float(fila[0]) if fila else float(TIPO_CAMBIO_INICIAL)


def guardar_tipo_cambio(valor):
    if valor <= 0:
        raise ValueError("El tipo de cambio debe ser mayor que cero.")

    with conectar() as conn:
        conn.execute(
            """
            INSERT INTO configuracion (clave, valor, actualizado_en)
            VALUES ('tipo_cambio_usd_bs', ?, ?)
            ON CONFLICT(clave) DO UPDATE SET
                valor = excluded.valor,
                actualizado_en = excluded.actualizado_en
            """,
            (
                str(float(valor)),
                datetime.now().isoformat(timespec="seconds"),
            ),
        )
        conn.commit()


def etiqueta_curso(codigo):
    try:
        datos = obtener_curso(codigo)
        if datos["nombre"] and datos["nombre"] != codigo:
            return f"{codigo} · {datos['nombre']}"
    except KeyError:
        pass
    return codigo


def codigo_desde_etiqueta(etiqueta):
    return etiqueta.split(" · ", 1)[0].strip()

def crear_hash_password(password, salt_hex=None):
    if salt_hex is None:
        salt = secrets.token_bytes(32)
        salt_hex = salt.hex()
    else:
        salt = bytes.fromhex(salt_hex)

    resultado = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        200_000,
    )
    return resultado.hex(), salt_hex


def verificar_password(password, hash_guardado, salt_hex):
    hash_ingresado, _ = crear_hash_password(password, salt_hex)
    return hmac.compare_digest(hash_ingresado, hash_guardado)


def crear_usuario(nombre, usuario, password, rol):
    nombre = nombre.strip()
    usuario = usuario.strip().lower()

    if not nombre or not usuario or not password:
        raise ValueError("Nombre, usuario y contraseña son obligatorios.")
    if rol not in ROLES:
        raise ValueError("El rol seleccionado no es válido.")
    if len(password) < 8:
        raise ValueError("La contraseña debe tener al menos 8 caracteres.")

    password_hash, salt = crear_hash_password(password)

    try:
        with conectar() as conn:
            conn.execute(
                """
                INSERT INTO usuarios (
                    nombre, usuario, password_hash, salt, rol, activo, creado_en
                )
                VALUES (?, ?, ?, ?, ?, 1, ?)
                """,
                (
                    nombre,
                    usuario,
                    password_hash,
                    salt,
                    rol,
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )
            conn.commit()
    except sqlite3.IntegrityError as error:
        raise ValueError("Ese nombre de usuario ya existe.") from error


def asegurar_administrador_inicial():
    with conectar() as conn:
        cantidad = conn.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0]

    if cantidad == 0:
        crear_usuario(
            "Administrador principal",
            "admin",
            "Admin123!",
            "Administrador",
        )


def autenticar_usuario(usuario, password):
    usuario = usuario.strip().lower()
    with conectar() as conn:
        fila = conn.execute(
            """
            SELECT id, nombre, usuario, password_hash, salt, rol, activo
            FROM usuarios
            WHERE usuario = ?
            """,
            (usuario,),
        ).fetchone()

    if fila is None:
        return None

    id_usuario, nombre, usuario_db, hash_guardado, salt, rol, activo = fila
    if not activo or not verificar_password(password, hash_guardado, salt):
        return None

    return {
        "id": id_usuario,
        "nombre": nombre,
        "usuario": usuario_db,
        "rol": rol,
    }


def cargar_usuarios():
    with conectar() as conn:
        return pd.read_sql_query(
            """
            SELECT id, nombre, usuario, rol, activo, creado_en
            FROM usuarios
            ORDER BY nombre
            """,
            conn,
        )


def cambiar_estado_usuario(id_usuario, activo):
    with conectar() as conn:
        conn.execute(
            "UPDATE usuarios SET activo = ? WHERE id = ?",
            (1 if activo else 0, int(id_usuario)),
        )
        conn.commit()


def cambiar_password_usuario(id_usuario, nueva_password):
    if len(nueva_password) < 8:
        raise ValueError("La contraseña debe tener al menos 8 caracteres.")

    password_hash, salt = crear_hash_password(nueva_password)
    with conectar() as conn:
        conn.execute(
            "UPDATE usuarios SET password_hash = ?, salt = ? WHERE id = ?",
            (password_hash, salt, int(id_usuario)),
        )
        conn.commit()


def numero_transaccion_existe(numero_transaccion):
    numero = numero_transaccion.strip()
    if not numero:
        return False

    with conectar() as conn:
        resultado = conn.execute(
            """
            SELECT 1
            FROM ventas
            WHERE LOWER(TRIM(numero_transaccion)) = LOWER(TRIM(?))
            LIMIT 1
            """,
            (numero,),
        ).fetchone()

    return resultado is not None


def guardar_venta(
    fecha,
    curso,
    monto,
    cliente,
    vendedor,
    sede,
    tipo_cambio,
    creado_por,
    numero_transaccion,
    banco,
    fecha_hora_pago,
    estado_pago,
    comprobante_bytes,
    comprobante_mime,
    comprobante_nombre,
):
    datos = COMISIONES[curso]
    comision_bs = (
        datos["monto"] * tipo_cambio
        if datos["moneda"] == "USD"
        else datos["monto"]
    )
    ingreso_neto = monto - comision_bs

    with conectar() as conn:
        conn.execute(
            """
            INSERT INTO ventas (
                fecha, curso, monto, cliente, vendedor, sede,
                comision_original, moneda_comision, tipo_cambio,
                comision_bs, ingreso_neto_bs, creado_por, creado_en,
                numero_transaccion, banco, fecha_hora_pago, estado_pago,
                comprobante, comprobante_mime, comprobante_nombre
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                fecha.isoformat(),
                curso,
                float(monto),
                cliente.strip(),
                vendedor.strip(),
                sede.strip(),
                datos["monto"],
                datos["moneda"],
                float(tipo_cambio),
                float(comision_bs),
                float(ingreso_neto),
                creado_por,
                datetime.now().isoformat(timespec="seconds"),
                numero_transaccion.strip(),
                banco.strip(),
                fecha_hora_pago.isoformat(timespec="minutes"),
                estado_pago,
                comprobante_bytes,
                comprobante_mime,
                comprobante_nombre,
            ),
        )
        conn.commit()

    return comision_bs, ingreso_neto

def cargar_ventas():
    with conectar() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM ventas ORDER BY fecha DESC, id DESC",
            conn,
        )

    if not df.empty:
        df["fecha"] = pd.to_datetime(df["fecha"])
        df["creado_en"] = pd.to_datetime(df["creado_en"])
    return df


def actualizar_venta(id_venta, fecha, curso, monto, cliente, vendedor, sede, tipo_cambio):
    datos = COMISIONES[curso]
    comision_bs = (
        datos["monto"] * tipo_cambio
        if datos["moneda"] == "USD"
        else datos["monto"]
    )
    ingreso_neto = monto - comision_bs

    with conectar() as conn:
        conn.execute(
            """
            UPDATE ventas
            SET fecha = ?, curso = ?, monto = ?, cliente = ?, vendedor = ?, sede = ?,
                comision_original = ?, moneda_comision = ?, tipo_cambio = ?,
                comision_bs = ?, ingreso_neto_bs = ?
            WHERE id = ?
            """,
            (
                fecha.isoformat(),
                curso,
                float(monto),
                cliente.strip(),
                vendedor.strip(),
                sede.strip(),
                datos["monto"],
                datos["moneda"],
                float(tipo_cambio),
                float(comision_bs),
                float(ingreso_neto),
                int(id_venta),
            ),
        )
        conn.commit()


def registrar_auditoria(usuario, accion, detalle=""):
    with conectar() as conn:
        conn.execute(
            """
            INSERT INTO auditoria_datos (usuario, accion, detalle, creado_en)
            VALUES (?, ?, ?, ?)
            """,
            (
                usuario,
                accion,
                detalle,
                datetime.now().isoformat(timespec="seconds"),
            ),
        )
        conn.commit()


def eliminar_venta(id_venta, usuario="sistema"):
    id_venta = int(id_venta)
    with conectar() as conn:
        fila = conn.execute(
            "SELECT cliente, numero_transaccion FROM ventas WHERE id = ?",
            (id_venta,),
        ).fetchone()
        if fila is None:
            raise ValueError("La venta seleccionada ya no existe.")

        # Conserva el registro del alumno, pero desvincula la venta eliminada.
        conn.execute(
            """
            UPDATE alumnos
            SET venta_id = NULL,
                estado = 'Con observación',
                observacion = CASE
                    WHEN TRIM(COALESCE(observacion, '')) = ''
                    THEN 'La venta vinculada fue eliminada por administración.'
                    ELSE observacion || ' | La venta vinculada fue eliminada por administración.'
                END
            WHERE venta_id = ?
            """,
            (id_venta,),
        )
        conn.execute("DELETE FROM ventas WHERE id = ?", (id_venta,))
        conn.execute(
            """
            INSERT INTO auditoria_datos (usuario, accion, detalle, creado_en)
            VALUES (?, 'ELIMINAR_VENTA', ?, ?)
            """,
            (
                usuario,
                f"Venta ID {id_venta}; cliente={fila[0]}; transacción={fila[1] or '-'}",
                datetime.now().isoformat(timespec="seconds"),
            ),
        )
        conn.commit()


def obtener_conteos_limpieza():
    with conectar() as conn:
        ventas = int(conn.execute("SELECT COUNT(*) FROM ventas").fetchone()[0])
        alumnos = int(conn.execute("SELECT COUNT(*) FROM alumnos").fetchone()[0])
    return ventas, alumnos


def limpiar_datos_operativos(eliminar_ventas, eliminar_alumnos, usuario):
    if not eliminar_ventas and not eliminar_alumnos:
        raise ValueError("Selecciona al menos un tipo de información para limpiar.")

    with conectar() as conn:
        conn.execute("BEGIN")
        ventas_antes = int(conn.execute("SELECT COUNT(*) FROM ventas").fetchone()[0])
        alumnos_antes = int(conn.execute("SELECT COUNT(*) FROM alumnos").fetchone()[0])

        # Si se eliminan ventas pero se conservan alumnos, los alumnos quedan observados.
        if eliminar_ventas and not eliminar_alumnos:
            conn.execute(
                """
                UPDATE alumnos
                SET venta_id = NULL,
                    estado = 'Con observación',
                    observacion = CASE
                        WHEN TRIM(COALESCE(observacion, '')) = ''
                        THEN 'Las ventas fueron limpiadas por administración.'
                        ELSE observacion || ' | Las ventas fueron limpiadas por administración.'
                    END
                WHERE venta_id IS NOT NULL
                """
            )

        if eliminar_alumnos:
            conn.execute("DELETE FROM alumnos")
            conn.execute("DELETE FROM sqlite_sequence WHERE name = 'alumnos'")

        if eliminar_ventas:
            conn.execute("DELETE FROM ventas")
            conn.execute("DELETE FROM sqlite_sequence WHERE name = 'ventas'")

        detalle = (
            f"Ventas eliminadas={ventas_antes if eliminar_ventas else 0}; "
            f"alumnos eliminados={alumnos_antes if eliminar_alumnos else 0}"
        )
        conn.execute(
            """
            INSERT INTO auditoria_datos (usuario, accion, detalle, creado_en)
            VALUES (?, 'LIMPIEZA_DATOS', ?, ?)
            """,
            (usuario, detalle, datetime.now().isoformat(timespec="seconds")),
        )
        conn.commit()

    return (
        ventas_antes if eliminar_ventas else 0,
        alumnos_antes if eliminar_alumnos else 0,
    )


def cargar_auditoria(limite=50):
    with conectar() as conn:
        return pd.read_sql_query(
            """
            SELECT usuario, accion, detalle, creado_en
            FROM auditoria_datos
            ORDER BY id DESC
            LIMIT ?
            """,
            conn,
            params=(int(limite),),
        )


def importar_dataframe(df, tipo_cambio):
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()
    df = df.loc[:, ~df.columns.str.contains("^Unnamed", case=False, regex=True)]

    if "Tipo_ingreso" in df.columns and "Curso" not in df.columns:
        df = df.rename(columns={"Tipo_ingreso": "Curso"})

    obligatorias = ["Fecha", "Curso", "Monto", "Cliente", "Vendedor"]
    faltantes = [c for c in obligatorias if c not in df.columns]
    if faltantes:
        raise ValueError(f"Faltan columnas obligatorias: {faltantes}")

    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce", dayfirst=True)
    df["Curso"] = df["Curso"].astype(str).str.strip().str.upper()
    df["Monto"] = (
        df["Monto"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("Bs", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.strip()
    )
    df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce")
    df["Cliente"] = df["Cliente"].astype(str).str.strip()
    df["Vendedor"] = df["Vendedor"].astype(str).str.strip()
    if "Sede" not in df.columns:
        df["Sede"] = "Sin especificar"
    df["Sede"] = df["Sede"].astype(str).str.strip().replace("", "Sin especificar")

    errores = df[
        df["Fecha"].isna()
        | df["Monto"].isna()
        | (df["Monto"] <= 0)
        | ~df["Curso"].isin(COMISIONES.keys())
        | (df["Cliente"] == "")
        | (df["Vendedor"] == "")
    ].copy()

    validos = df.drop(index=errores.index).copy()

    registros = []
    for _, fila in validos.iterrows():
        curso = fila["Curso"]
        datos = COMISIONES[curso]
        comision_bs = (
            datos["monto"] * tipo_cambio
            if datos["moneda"] == "USD"
            else datos["monto"]
        )
        ingreso_neto = fila["Monto"] - comision_bs

        registros.append(
            (
                fila["Fecha"].date().isoformat(),
                curso,
                float(fila["Monto"]),
                fila["Cliente"],
                fila["Vendedor"],
                fila["Sede"],
                datos["monto"],
                datos["moneda"],
                float(tipo_cambio),
                float(comision_bs),
                float(ingreso_neto),
                datetime.now().isoformat(timespec="seconds"),
            )
        )

    if registros:
        with conectar() as conn:
            conn.executemany(
                """
                INSERT INTO ventas (
                    fecha, curso, monto, cliente, vendedor, sede,
                    comision_original, moneda_comision, tipo_cambio,
                    comision_bs, ingreso_neto_bs, creado_en
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                registros,
            )
            conn.commit()

    return len(validos), errores


# ============================================================
# REGISTRO Y CONTROL DE ALUMNOS
# ============================================================

def normalizar_transaccion(valor):
    return re.sub(r"\s+", "", (valor or "").strip()).lower()


def buscar_venta_por_transaccion(numero_transaccion):
    numero = (numero_transaccion or "").strip()
    if not numero:
        return None

    with conectar() as conn:
        fila = conn.execute(
            """
            SELECT id, cliente, curso, sede, monto, vendedor, banco,
                   fecha_hora_pago, estado_pago, numero_transaccion
            FROM ventas
            WHERE LOWER(REPLACE(TRIM(numero_transaccion), ' ', '')) = ?
            LIMIT 1
            """,
            (normalizar_transaccion(numero),),
        ).fetchone()

    if fila is None:
        return None

    columnas = [
        "id", "cliente", "curso", "sede", "monto", "vendedor", "banco",
        "fecha_hora_pago", "estado_pago", "numero_transaccion",
    ]
    return dict(zip(columnas, fila))


def codigo_alumno_nuevo():
    prefijo = f"JAA-{datetime.now().year}-"
    with conectar() as conn:
        ultimo = conn.execute(
            "SELECT MAX(id) FROM alumnos"
        ).fetchone()[0] or 0
    return f"{prefijo}{ultimo + 1:05d}"


def alumno_ya_registrado(numero_transaccion, documento):
    with conectar() as conn:
        fila = conn.execute(
            """
            SELECT codigo
            FROM alumnos
            WHERE LOWER(REPLACE(TRIM(numero_transaccion), ' ', '')) = ?
               OR LOWER(TRIM(documento)) = LOWER(TRIM(?))
            LIMIT 1
            """,
            (normalizar_transaccion(numero_transaccion), documento.strip()),
        ).fetchone()
    return fila[0] if fila else None


def guardar_alumno(
    nombre_completo, documento, telefono, correo, numero_transaccion,
    curso_manual=None, sede_manual=None, monto_reportado=None,
):
    nombre_completo = nombre_completo.strip()
    documento = documento.strip()
    telefono = telefono.strip()
    correo = correo.strip()
    numero_transaccion = numero_transaccion.strip()

    if not nombre_completo or not documento or not telefono or not numero_transaccion:
        raise ValueError(
            "Nombre completo, documento, teléfono y número de transacción son obligatorios."
        )

    existente = alumno_ya_registrado(numero_transaccion, documento)
    if existente:
        raise ValueError(
            f"Ya existe un registro relacionado con estos datos: {existente}."
        )

    venta = buscar_venta_por_transaccion(numero_transaccion)
    if venta:
        venta_id = int(venta["id"])
        curso = venta["curso"]
        sede = venta["sede"]
        monto = float(venta["monto"])
        estado = "Validado"
        observacion = "Venta encontrada automáticamente por número de transacción."
    else:
        venta_id = None
        curso = (curso_manual or "").strip()
        sede = (sede_manual or "").strip()
        monto = float(monto_reportado or 0)
        estado = "Con observación"
        observacion = "No se encontró una venta registrada con este número de transacción."

    codigo = codigo_alumno_nuevo()
    with conectar() as conn:
        conn.execute(
            """
            INSERT INTO alumnos (
                codigo, nombre_completo, documento, telefono, correo,
                numero_transaccion, venta_id, curso, sede, monto_reportado,
                estado, observacion, registrado_en
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                codigo, nombre_completo, documento, telefono, correo,
                numero_transaccion, venta_id, curso, sede, monto, estado,
                observacion, datetime.now().isoformat(timespec="seconds"),
            ),
        )
        conn.commit()
    return codigo, estado, venta


def cargar_alumnos():
    with conectar() as conn:
        return pd.read_sql_query(
            """
            SELECT a.*, v.cliente AS cliente_venta, v.vendedor, v.banco,
                   v.estado_pago, v.monto AS monto_venta
            FROM alumnos a
            LEFT JOIN ventas v ON v.id = a.venta_id
            ORDER BY a.registrado_en DESC, a.id DESC
            """,
            conn,
        )


def actualizar_estado_alumno(id_alumno, estado, observacion, revisado_por):
    estados_validos = {"Pendiente", "Validado", "Con observación", "Rechazado"}
    if estado not in estados_validos:
        raise ValueError("Estado de alumno no válido.")

    with conectar() as conn:
        conn.execute(
            """
            UPDATE alumnos
            SET estado = ?, observacion = ?, revisado_por = ?, revisado_en = ?
            WHERE id = ?
            """,
            (
                estado, observacion.strip(), revisado_por,
                datetime.now().isoformat(timespec="seconds"), int(id_alumno),
            ),
        )
        conn.commit()


def intentar_vincular_alumno(id_alumno, revisado_por):
    with conectar() as conn:
        fila = conn.execute(
            "SELECT numero_transaccion FROM alumnos WHERE id = ?",
            (int(id_alumno),),
        ).fetchone()
    if fila is None:
        raise ValueError("El alumno seleccionado ya no existe.")

    venta = buscar_venta_por_transaccion(fila[0])
    if venta is None:
        raise ValueError("Todavía no existe una venta con esa transacción.")

    with conectar() as conn:
        conn.execute(
            """
            UPDATE alumnos
            SET venta_id = ?, curso = ?, sede = ?, monto_reportado = ?,
                estado = 'Validado',
                observacion = 'Venta vinculada posteriormente por administración.',
                revisado_por = ?, revisado_en = ?
            WHERE id = ?
            """,
            (
                int(venta["id"]), venta["curso"], venta["sede"],
                float(venta["monto"]), revisado_por,
                datetime.now().isoformat(timespec="seconds"), int(id_alumno),
            ),
        )
        conn.commit()


# ============================================================
# REPORTE EXCEL
# ============================================================

def generar_reporte_excel(df):
    if df.empty:
        raise ValueError("No existen ventas para exportar.")

    trabajo = df.copy()
    trabajo["Fecha"] = trabajo["fecha"].dt.date
    trabajo["Mes"] = trabajo["fecha"].dt.to_period("M").astype(str)

    ventas = trabajo[
        [
            "id",
            "Fecha",
            "curso",
            "monto",
            "cliente",
            "vendedor",
            "sede",
            "comision_original",
            "moneda_comision",
            "tipo_cambio",
            "comision_bs",
            "ingreso_neto_bs",
            "numero_transaccion",
            "banco",
            "fecha_hora_pago",
            "estado_pago",
            "comprobante_nombre",
        ]
    ].rename(
        columns={
            "id": "ID",
            "curso": "Curso",
            "monto": "Ingreso_bruto_Bs",
            "cliente": "Cliente",
            "vendedor": "Vendedor",
            "sede": "Sede",
            "comision_original": "Comision_original",
            "moneda_comision": "Moneda_comision",
            "tipo_cambio": "Tipo_cambio",
            "comision_bs": "Comision_Bs",
            "ingreso_neto_bs": "Ingreso_neto_Bs",
            "numero_transaccion": "Numero_transaccion",
            "banco": "Banco",
            "fecha_hora_pago": "Fecha_hora_pago",
            "estado_pago": "Estado_pago",
            "comprobante_nombre": "Comprobante",
        }
    )

    resumen_curso = (
        trabajo.groupby("curso", as_index=False)
        .agg(
            Ingreso_bruto_Bs=("monto", "sum"),
            Comisiones_Bs=("comision_bs", "sum"),
            Ingreso_neto_Bs=("ingreso_neto_bs", "sum"),
            Cantidad_ventas=("id", "count"),
            Venta_promedio_Bs=("monto", "mean"),
        )
        .rename(columns={"curso": "Curso"})
        .sort_values("Ingreso_neto_Bs", ascending=False)
    )

    resumen_vendedor = (
        trabajo.groupby("vendedor", as_index=False)
        .agg(
            Ingreso_bruto_Bs=("monto", "sum"),
            Comisiones_Bs=("comision_bs", "sum"),
            Ingreso_neto_Bs=("ingreso_neto_bs", "sum"),
            Cantidad_ventas=("id", "count"),
        )
        .rename(columns={"vendedor": "Vendedor"})
        .sort_values("Ingreso_neto_Bs", ascending=False)
    )

    resumen_sede = (
        trabajo.assign(
            sede=trabajo["sede"].fillna("").replace("", "Sin especificar")
        )
        .groupby("sede", as_index=False)
        .agg(
            Ingreso_bruto_Bs=("monto", "sum"),
            Comisiones_Bs=("comision_bs", "sum"),
            Ingreso_neto_Bs=("ingreso_neto_bs", "sum"),
            Cantidad_ventas=("id", "count"),
        )
        .rename(columns={"sede": "Sede"})
        .sort_values("Ingreso_neto_Bs", ascending=False)
    )

    resumen_diario = (
        trabajo.groupby(["Fecha", "curso"], as_index=False)
        .agg(
            Ingreso_bruto_Bs=("monto", "sum"),
            Comisiones_Bs=("comision_bs", "sum"),
            Ingreso_neto_Bs=("ingreso_neto_bs", "sum"),
            Cantidad_ventas=("id", "count"),
        )
        .rename(columns={"curso": "Curso"})
        .sort_values(["Fecha", "Curso"])
    )

    resumen_mensual = (
        trabajo.groupby(["Mes", "curso"], as_index=False)
        .agg(
            Ingreso_bruto_Bs=("monto", "sum"),
            Comisiones_Bs=("comision_bs", "sum"),
            Ingreso_neto_Bs=("ingreso_neto_bs", "sum"),
            Cantidad_ventas=("id", "count"),
        )
        .rename(columns={"curso": "Curso"})
        .sort_values(["Mes", "Curso"])
    )

    curso_top = resumen_curso.iloc[0]["Curso"]
    vendedor_top = resumen_vendedor.iloc[0]["Vendedor"]

    dashboard = pd.DataFrame(
        {
            "Indicador": [
                "Ingreso bruto total",
                "Comisiones totales",
                "Ingreso neto total",
                "Cantidad de ventas",
                "Ticket promedio",
                "Curso con mayor ingreso neto",
                "Vendedor con mayor ingreso neto",
            ],
            "Valor": [
                trabajo["monto"].sum(),
                trabajo["comision_bs"].sum(),
                trabajo["ingreso_neto_bs"].sum(),
                len(trabajo),
                trabajo["monto"].mean(),
                curso_top,
                vendedor_top,
            ],
        }
    )

    salida = BytesIO()

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        dashboard.to_excel(writer, sheet_name="Dashboard", index=False, startrow=1)
        ventas.to_excel(writer, sheet_name="Ventas", index=False)
        resumen_curso.to_excel(writer, sheet_name="Resumen por curso", index=False)
        resumen_vendedor.to_excel(writer, sheet_name="Resumen vendedores", index=False)
        resumen_sede.to_excel(writer, sheet_name="Resumen sedes", index=False)
        resumen_diario.to_excel(writer, sheet_name="Resumen diario", index=False)
        resumen_mensual.to_excel(writer, sheet_name="Resumen mensual", index=False)

        nombres_hojas = set(writer.book.sheetnames)
        for curso in sorted(ventas["Curso"].dropna().astype(str).unique()):
            nombre_hoja = re.sub(r'[\\/*?:\[\]]', '_', curso)[:31] or "Curso"
            base_hoja = nombre_hoja
            contador = 2
            while nombre_hoja in nombres_hojas:
                sufijo = f"_{contador}"
                nombre_hoja = f"{base_hoja[:31-len(sufijo)]}{sufijo}"
                contador += 1
            nombres_hojas.add(nombre_hoja)
            ventas[ventas["Curso"] == curso].to_excel(
                writer,
                sheet_name=nombre_hoja,
                index=False,
            )

    salida.seek(0)
    wb = load_workbook(salida)

    azul = "1F4E78"
    azul_claro = "D9EAF7"
    blanco = "FFFFFF"
    borde = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for ws in wb.worksheets:
        fila_encabezado = 2 if ws.title == "Dashboard" else 1
        ws.freeze_panes = f"A{fila_encabezado + 1}"
        ws.auto_filter.ref = ws.dimensions

        for celda in ws[fila_encabezado]:
            celda.fill = PatternFill("solid", fgColor=azul)
            celda.font = Font(color=blanco, bold=True)
            celda.alignment = Alignment(horizontal="center")
            celda.border = borde

        for fila in ws.iter_rows():
            for celda in fila:
                celda.border = borde

        for columna in ws.columns:
            letra = get_column_letter(columna[0].column)
            ancho = max(
                len(str(celda.value)) if celda.value is not None else 0
                for celda in columna
            )
            ws.column_dimensions[letra].width = min(ancho + 3, 34)

        for fila in ws.iter_rows(min_row=fila_encabezado + 1):
            for celda in fila:
                encabezado = ws.cell(row=fila_encabezado, column=celda.column).value
                if encabezado and any(
                    palabra in str(encabezado)
                    for palabra in ["Monto", "Ingreso", "Comision", "Valor", "Venta"]
                ):
                    if isinstance(celda.value, (int, float)):
                        celda.number_format = '#,##0.00'

    ws_dashboard = wb["Dashboard"]
    ws_dashboard["A1"] = "DASHBOARD EJECUTIVO DE VENTAS"
    ws_dashboard.merge_cells("A1:B1")
    ws_dashboard["A1"].font = Font(size=16, bold=True, color=blanco)
    ws_dashboard["A1"].fill = PatternFill("solid", fgColor=azul)
    ws_dashboard["A1"].alignment = Alignment(horizontal="center")

    for fila in range(3, ws_dashboard.max_row + 1):
        ws_dashboard[f"A{fila}"].fill = PatternFill("solid", fgColor=azul_claro)
        ws_dashboard[f"A{fila}"].font = Font(bold=True)

    ws_curso = wb["Resumen por curso"]
    grafico = BarChart()
    grafico.title = "Ingreso neto por curso"
    grafico.y_axis.title = "Bolivianos"
    grafico.x_axis.title = "Curso"
    datos = Reference(ws_curso, min_col=4, min_row=1, max_row=ws_curso.max_row)
    categorias = Reference(ws_curso, min_col=1, min_row=2, max_row=ws_curso.max_row)
    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.height = 8
    grafico.width = 14
    ws_curso.add_chart(grafico, "G2")

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()


# ============================================================
# INTERFAZ
# ============================================================

crear_base_datos()
asegurar_administrador_inicial()

TIPO_CAMBIO_USD_BS = obtener_tipo_cambio()
SEDES = cargar_catalogo("sedes", solo_activos=True)["nombre"].tolist()
BANCOS = cargar_catalogo("bancos", solo_activos=True)["nombre"].tolist()

if "usuario_actual" not in st.session_state:
    st.session_state.usuario_actual = None


def cerrar_sesion():
    st.session_state.usuario_actual = None
    st.rerun()


# Enlace público: agrega ?registro=alumno al URL de la aplicación.
modo_registro_alumno = st.query_params.get("registro") == "alumno"

if modo_registro_alumno:
    st.markdown(
        """
        <div class="jet-hero">
            <div class="jet-kicker">Registro oficial de estudiantes</div>
            <div class="jet-hero-title">Confirma tu inscripción</div>
            <p class="jet-hero-copy">
                Después de realizar tu pago, completa este formulario.
                Tu información será comparada con la venta registrada por la academia.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.info(
        "Ten a mano el número de transacción de tu pago. "
        "No necesitas crear una cuenta ni iniciar sesión."
    )

    numero_publico = st.text_input(
        "Número de transacción",
        placeholder="Ej.: 45893271 o TRX-45893271",
        key="transaccion_alumno_publica",
    )

    if numero_publico.strip():
        venta_publica = buscar_venta_por_transaccion(numero_publico)
        if venta_publica:
            st.success("Venta encontrada en el sistema.")
            d1, d2, d3 = st.columns(3)
            d1.metric("Curso", venta_publica["curso"] or "-")
            d2.metric("Sede", venta_publica["sede"] or "-")
            d3.metric("Monto", f"Bs {float(venta_publica['monto']):,.2f}")
            st.caption(
                f"Vendedor registrado: {venta_publica['vendedor'] or '-'} · "
                f"Estado del pago: {venta_publica['estado_pago'] or 'Pendiente'}"
            )
        else:
            st.warning(
                "No encontramos esa transacción. Puedes registrarte igualmente; "
                "tu solicitud quedará con observación para que administración la revise."
            )

        with st.form("registro_publico_alumno", clear_on_submit=False):
            a1, a2 = st.columns(2)
            with a1:
                nombre_alumno = st.text_input("Nombre completo")
                documento_alumno = st.text_input("CI o pasaporte")
                telefono_alumno = st.text_input("Celular / WhatsApp")
            with a2:
                correo_alumno = st.text_input("Correo electrónico (opcional)")
                if venta_publica:
                    curso_alumno = venta_publica["curso"]
                    sede_alumno = venta_publica["sede"]
                    monto_alumno = float(venta_publica["monto"])
                    st.text_input("Curso", value=curso_alumno or "", disabled=True)
                    st.text_input("Sede", value=sede_alumno or "", disabled=True)
                else:
                    opciones_publicas = [etiqueta_curso(c) for c in COMISIONES.keys()]
                    curso_publico = st.selectbox("Curso pagado", opciones_publicas)
                    curso_alumno = codigo_desde_etiqueta(curso_publico)
                    sede_alumno = st.selectbox("Sede", SEDES)
                    monto_alumno = st.number_input(
                        "Monto pagado (Bs)", min_value=0.0, value=0.0, step=10.0
                    )

            aceptar = st.checkbox(
                "Confirmo que los datos ingresados son correctos y autorizo su revisión."
            )
            enviar_alumno = st.form_submit_button(
                "Enviar registro de inscripción",
                type="primary",
                use_container_width=True,
            )

            if enviar_alumno:
                if not aceptar:
                    st.error("Debes confirmar que los datos son correctos.")
                else:
                    try:
                        codigo, estado, _ = guardar_alumno(
                            nombre_alumno, documento_alumno, telefono_alumno,
                            correo_alumno, numero_publico, curso_alumno,
                            sede_alumno, monto_alumno,
                        )
                        st.success(
                            f"Registro enviado correctamente. Tu código es {codigo}."
                        )
                        if estado == "Validado":
                            st.success("Tu inscripción quedó validada automáticamente.")
                        else:
                            st.warning(
                                "Tu registro quedó con observación. Administración "
                                "revisará el pago y la transacción."
                            )
                    except ValueError as error:
                        st.error(str(error))

    st.divider()
    st.caption("Jet Airways Academy · Registro seguro de estudiantes")
    if st.button("Ir al acceso del personal"):
        st.query_params.clear()
        st.rerun()
    st.stop()


if st.session_state.usuario_actual is None:
    izquierda, centro, derecha = st.columns([1, 1.05, 1])

    with centro:
        if LOGO_PATH.exists():
            logo_col_1, logo_col_2, logo_col_3 = st.columns([0.45, 1, 0.45])
            with logo_col_2:
                st.image(str(LOGO_PATH), use_container_width=True)

        st.markdown(
            '<div class="jet-brand-title">Sistema de Gestión de Ventas</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div class="jet-brand-subtitle">'
            'Jet Airways Academy · Acceso exclusivo para usuarios autorizados'
            '</div>',
            unsafe_allow_html=True,
        )

        with st.form("formulario_login"):
            usuario_login = st.text_input(
                "Usuario",
                placeholder="Ingresa tu nombre de usuario",
            )
            password_login = st.text_input(
                "Contraseña",
                type="password",
                placeholder="Ingresa tu contraseña",
            )
            entrar = st.form_submit_button(
                "Iniciar sesión",
                type="primary",
                use_container_width=True,
            )

        if entrar:
            datos_usuario = autenticar_usuario(usuario_login, password_login)
            if datos_usuario is None:
                st.error("Usuario o contraseña incorrectos.")
            else:
                st.session_state.usuario_actual = datos_usuario
                st.rerun()

        with st.expander("Acceso inicial"):
            st.write("Usuario: `admin`")
            st.write("Contraseña: `Admin123!`")
            st.warning("Cambia esta contraseña después del primer ingreso.")

        st.markdown(
            '<div class="jet-footer">'
            'Jet Airways Academy · Sistema interno de gestión © 2026'
            '</div>',
            unsafe_allow_html=True,
        )

    st.stop()

usuario_actual = st.session_state.usuario_actual
rol_actual = usuario_actual["rol"]

cabecera_1, cabecera_2 = st.columns([0.15, 0.85])
with cabecera_1:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), use_container_width=True)
with cabecera_2:
    st.title("Jet Airways Academy")
    st.caption("Executive Sales CRM · Plataforma interna de gestión comercial")
st.caption(
    f"Sesión: **{usuario_actual['nombre']}** · "
    f"Usuario: **{usuario_actual['usuario']}** · "
    f"Rol: **{rol_actual}**"
)

st.markdown(
    f"""
    <div class="jet-hero">
        <div class="jet-kicker">Executive Sales Intelligence</div>
        <div class="jet-hero-title">Centro de Gestión Comercial</div>
        <p class="jet-hero-copy">
            Ventas, comisiones, comprobantes y rendimiento comercial reunidos
            en una sola plataforma segura para Jet Airways Academy.
        </p>
        <span class="jet-badge">
            Sesión activa · {usuario_actual["nombre"]} · {rol_actual}
        </span>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    if LOGO_PATH.exists():
        logo_col_1, logo_col_2, logo_col_3 = st.columns([0.25, 1, 0.25])
        with logo_col_2:
            st.image(str(LOGO_PATH), use_container_width=True)
    else:
        st.markdown(
            """
            <div class="sidebar-brand">
                <div class="sidebar-brand-icon">✈️</div>
                <div class="sidebar-brand-name">Jet Airways Academy</div>
                <div class="sidebar-brand-sub">Executive Sales CRM</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown(
        '<div class="jet-section-title">Navegación</div>',
        unsafe_allow_html=True,
    )

    etiquetas_menu = {
        "Registrar venta": "✦  Registrar venta",
        "Dashboard": "◈  Dashboard ejecutivo",
        "Consultar ventas": "⌕  Consultar ventas",
        "Alumnos": "🎓  Control de alumnos",
        "Exportar y respaldo": "⇩  Exportar y respaldo",
        "Editar o eliminar": "✎  Editar o eliminar",
        "Limpieza de datos": "🧹  Limpieza de datos",
        "Importar Excel": "↥  Importar Excel",
        "Usuarios": "◎  Usuarios",
        "Configuración": "⚙  Configuración",
    }

    opciones_internas = ["Registrar venta", "Dashboard", "Consultar ventas"]

    if rol_actual in ["Administrador", "Supervisor"]:
        opciones_internas.extend(["Alumnos", "Exportar y respaldo"])

    if rol_actual == "Administrador":
        opciones_internas.extend(
            [
                "Editar o eliminar",
                "Limpieza de datos",
                "Importar Excel",
                "Usuarios",
                "Configuración",
            ]
        )

    opciones_visibles = [etiquetas_menu[opcion] for opcion in opciones_internas]
    seleccion_visible = st.radio(
        "Selecciona una sección",
        opciones_visibles,
        label_visibility="collapsed",
    )
    seccion = {
        etiqueta: interna
        for interna, etiqueta in etiquetas_menu.items()
    }[seleccion_visible]
    st.divider()
    st.write(f"**{usuario_actual['nombre']}**")
    st.caption(rol_actual)
    if st.button("Cerrar sesión", use_container_width=True):
        cerrar_sesion()

# ------------------------------------------------------------
# REGISTRAR
# ------------------------------------------------------------

if seccion == "Registrar venta":
    st.markdown('<div class="jet-section-title">Registrar nueva venta</div>', unsafe_allow_html=True)
    st.caption(
        "Registra la información de la venta y adjunta el extracto o comprobante de pago."
    )

    with st.form("registro_venta", clear_on_submit=True):
        c1, c2 = st.columns(2)

        with c1:
            fecha = st.date_input("Fecha de la venta", value=date.today())
            opciones_cursos = [
                etiqueta_curso(codigo) for codigo in COMISIONES.keys()
            ]
            curso_etiqueta = st.selectbox("Curso", opciones_cursos)
            curso = codigo_desde_etiqueta(curso_etiqueta)
            monto = st.number_input(
                "Monto de la venta (Bs)",
                min_value=0.01,
                value=1000.00,
                step=10.00,
            )
            cliente = st.text_input("Cliente")

        with c2:
            sede = st.selectbox(
                "Sede",
                SEDES,
            )
            vendedor = st.text_input(
                "Nombre del vendedor",
                placeholder="Escribe tu nombre completo",
                help="Escribe el nombre tal como debe aparecer en los reportes.",
            )

            tipo_cambio = st.number_input(
                "Tipo de cambio USD/BOB",
                min_value=0.01,
                value=float(TIPO_CAMBIO_USD_BS),
                step=0.01,
            )
            numero_transaccion = st.text_input(
                "Número de transacción",
                placeholder="Ej.: 45893271 o TRX-45893271",
            )
            banco = st.selectbox(
                "Banco o medio de pago",
                BANCOS,
            )

        st.write("#### Información del pago")
        p1, p2 = st.columns(2)
        with p1:
            fecha_pago = st.date_input("Fecha del pago", value=date.today())
        with p2:
            hora_pago = st.time_input(
                "Hora del pago",
                value=datetime.now().time().replace(second=0, microsecond=0),
            )

        estado_pago = st.selectbox(
            "Estado del pago",
            ["Pendiente", "Verificando", "Aprobado", "Rechazado"],
            index=0,
        )

        metodo_comprobante = st.radio(
            "¿Cómo deseas adjuntar el comprobante?",
            ["Subir archivo", "Tomar foto"],
            horizontal=True,
        )

        comprobante = None
        if metodo_comprobante == "Subir archivo":
            comprobante = st.file_uploader(
                "Sube el extracto o comprobante",
                type=["png", "jpg", "jpeg", "webp", "pdf"],
                help="Formatos permitidos: PNG, JPG, WEBP o PDF.",
            )
        else:
            comprobante = st.camera_input(
                "Toma una foto del extracto o comprobante"
            )

        datos_comision = COMISIONES[curso]
        comision_estimada = (
            datos_comision["monto"] * tipo_cambio
            if datos_comision["moneda"] == "USD"
            else datos_comision["monto"]
        )
        neto_estimado = monto - comision_estimada

        m1, m2, m3 = st.columns(3)
        m1.metric("Ingreso bruto", f"Bs {monto:,.2f}")
        m2.metric("Comisión", f"Bs {comision_estimada:,.2f}")
        m3.metric("Ingreso neto", f"Bs {neto_estimado:,.2f}")

        guardar = st.form_submit_button(
            "Guardar venta y comprobante",
            type="primary",
            use_container_width=True,
        )

        if guardar:
            vendedor_final = vendedor.strip()
            transaccion_limpia = numero_transaccion.strip()
            fecha_hora_pago = datetime.combine(fecha_pago, hora_pago)

            if not cliente.strip():
                st.error("Debes ingresar el nombre del cliente.")
            elif not vendedor_final:
                st.error("Debes ingresar el nombre del vendedor.")
            elif not transaccion_limpia:
                st.error("Debes ingresar el número de transacción.")
            elif numero_transaccion_existe(transaccion_limpia):
                st.error(
                    "Ese número de transacción ya está registrado. "
                    "Verifica el dato antes de continuar."
                )
            elif comprobante is None:
                st.error("Debes subir o tomar una foto del comprobante.")
            else:
                comprobante_bytes = comprobante.getvalue()
                comprobante_mime = getattr(
                    comprobante,
                    "type",
                    "image/jpeg",
                )
                comprobante_nombre = getattr(
                    comprobante,
                    "name",
                    f"comprobante_{transaccion_limpia}.jpg",
                )

                try:
                    comision, neto = guardar_venta(
                        fecha,
                        curso,
                        monto,
                        cliente,
                        vendedor_final,
                        sede,
                        tipo_cambio,
                        usuario_actual["usuario"],
                        transaccion_limpia,
                        banco,
                        fecha_hora_pago,
                        estado_pago,
                        comprobante_bytes,
                        comprobante_mime,
                        comprobante_nombre,
                    )
                    st.success(
                        f"Venta guardada correctamente. "
                        f"Comisión: Bs {comision:,.2f} | "
                        f"Ingreso neto: Bs {neto:,.2f}"
                    )
                    if neto < 0:
                        st.warning("La comisión supera el monto de la venta.")
                except sqlite3.IntegrityError:
                    st.error(
                        "No se pudo guardar: el número de transacción ya existe."
                    )


# ------------------------------------------------------------
# DASHBOARD
# ------------------------------------------------------------

elif seccion == "Dashboard":
    df = cargar_ventas()

    if rol_actual == "Vendedor" and not df.empty:
        df = df[df["creado_por"] == usuario_actual["usuario"]].copy()

    if df.empty:
        st.info("Todavía no existen ventas registradas.")
        st.stop()

    st.markdown('<div class="jet-section-title">Dashboard ejecutivo</div>', unsafe_allow_html=True)

    minimo = df["fecha"].min().date()
    maximo = df["fecha"].max().date()

    df["sede"] = df["sede"].fillna("").replace("", "Sin especificar")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        fechas = st.date_input(
            "Rango de fechas",
            value=(minimo, maximo),
            min_value=minimo,
            max_value=maximo,
        )
    with c2:
        sedes = st.multiselect(
            "Sedes",
            sorted(df["sede"].unique()),
            default=sorted(df["sede"].unique()),
        )
    with c3:
        cursos = st.multiselect(
            "Cursos",
            sorted(df["curso"].unique()),
            default=sorted(df["curso"].unique()),
        )
    with c4:
        vendedores = st.multiselect(
            "Vendedores",
            sorted(df["vendedor"].unique()),
            default=sorted(df["vendedor"].unique()),
        )

    if isinstance(fechas, tuple) and len(fechas) == 2:
        inicio, fin = fechas
    else:
        inicio = fin = fechas

    filtrado = df[
        (df["fecha"].dt.date >= inicio)
        & (df["fecha"].dt.date <= fin)
        & (df["sede"].isin(sedes))
        & (df["curso"].isin(cursos))
        & (df["vendedor"].isin(vendedores))
    ].copy()

    if filtrado.empty:
        st.warning("No existen datos para los filtros seleccionados.")
        st.stop()

    total_bruto = filtrado["monto"].sum()
    comisiones = filtrado["comision_bs"].sum()
    neto = filtrado["ingreso_neto_bs"].sum()
    cantidad = len(filtrado)
    promedio = filtrado["monto"].mean()

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Ingreso bruto", f"Bs {total_bruto:,.2f}")
    k2.metric("Comisiones", f"Bs {comisiones:,.2f}")
    k3.metric("Ingreso neto", f"Bs {neto:,.2f}")
    k4.metric("Ventas", f"{cantidad:,}")
    k5.metric("Ticket promedio", f"Bs {promedio:,.2f}")

    if "estado_pago" in filtrado.columns:
        estados = (
            filtrado["estado_pago"]
            .fillna("Pendiente")
            .replace("", "Pendiente")
            .value_counts()
        )
        st.markdown(
            '<div class="jet-section-title">Control de pagos</div>',
            unsafe_allow_html=True,
        )
        e1, e2, e3, e4 = st.columns(4)
        e1.metric("Pendientes", int(estados.get("Pendiente", 0)))
        e2.metric("Verificando", int(estados.get("Verificando", 0)))
        e3.metric("Aprobados", int(estados.get("Aprobado", 0)))
        e4.metric("Rechazados", int(estados.get("Rechazado", 0)))

    por_curso = (
        filtrado.groupby("curso", as_index=False)
        .agg(
            Ingreso_bruto=("monto", "sum"),
            Ingreso_neto=("ingreso_neto_bs", "sum"),
            Ventas=("id", "count"),
        )
        .sort_values("Ingreso_neto", ascending=False)
    )

    por_vendedor = (
        filtrado.groupby("vendedor", as_index=False)
        .agg(
            Ingreso_neto=("ingreso_neto_bs", "sum"),
            Ventas=("id", "count"),
        )
        .sort_values("Ingreso_neto", ascending=False)
    )

    g1, g2 = st.columns(2)

    with g1:
        fig = px.bar(
            por_curso,
            x="curso",
            y="Ingreso_neto",
            title="Ingreso neto por curso",
            labels={"curso": "Curso", "Ingreso_neto": "Ingreso neto (Bs)"},
        )
        st.plotly_chart(fig, use_container_width=True)

    with g2:
        fig = px.pie(
            por_curso,
            names="curso",
            values="Ingreso_neto",
            title="Participación del ingreso neto",
            hole=0.4,
        )
        st.plotly_chart(fig, use_container_width=True)

    tendencia = (
        filtrado.assign(Fecha=filtrado["fecha"].dt.date)
        .groupby("Fecha", as_index=False)["ingreso_neto_bs"]
        .sum()
    )

    fig = px.line(
        tendencia,
        x="Fecha",
        y="ingreso_neto_bs",
        markers=True,
        title="Evolución diaria del ingreso neto",
        labels={"ingreso_neto_bs": "Ingreso neto (Bs)"},
    )
    st.plotly_chart(fig, use_container_width=True)

    g3, g4 = st.columns(2)
    with g3:
        st.write("#### Ranking de cursos")
        st.dataframe(por_curso, use_container_width=True, hide_index=True)
    with g4:
        st.write("#### Ranking de vendedores")
        st.dataframe(por_vendedor, use_container_width=True, hide_index=True)

# ------------------------------------------------------------
# CONSULTAR
# ------------------------------------------------------------

elif seccion == "Consultar ventas":
    df = cargar_ventas()

    if rol_actual == "Vendedor" and not df.empty:
        df = df[df["creado_por"] == usuario_actual["usuario"]].copy()
    st.markdown('<div class="jet-section-title">Consultar y filtrar ventas</div>', unsafe_allow_html=True)

    if df.empty:
        st.info("No existen ventas registradas.")
        st.stop()

    c1, c2, c3 = st.columns(3)
    with c1:
        curso_filtro = st.selectbox(
            "Curso",
            ["Todos"] + sorted(df["curso"].unique().tolist()),
        )
    with c2:
        vendedor_filtro = st.selectbox(
            "Vendedor",
            ["Todos"] + sorted(df["vendedor"].unique().tolist()),
        )
    with c3:
        cliente_busqueda = st.text_input("Buscar cliente")

    filtrado = df.copy()

    if curso_filtro != "Todos":
        filtrado = filtrado[filtrado["curso"] == curso_filtro]

    if vendedor_filtro != "Todos":
        filtrado = filtrado[filtrado["vendedor"] == vendedor_filtro]

    if cliente_busqueda.strip():
        filtrado = filtrado[
            filtrado["cliente"].str.contains(
                cliente_busqueda.strip(),
                case=False,
                na=False,
            )
        ]

    tabla = filtrado.rename(
        columns={
            "id": "ID",
            "fecha": "Fecha",
            "curso": "Curso",
            "monto": "Monto",
            "cliente": "Cliente",
            "vendedor": "Vendedor",
            "comision_bs": "Comisión Bs",
            "ingreso_neto_bs": "Ingreso neto Bs",
            "creado_por": "Registrado por",
            "numero_transaccion": "N.º transacción",
            "banco": "Banco",
            "fecha_hora_pago": "Fecha/hora pago",
            "estado_pago": "Estado del pago",
            "comprobante_nombre": "Comprobante",
        }
    )[
        [
            "ID",
            "Fecha",
            "Curso",
            "Monto",
            "Cliente",
            "Vendedor",
            "Comisión Bs",
            "Ingreso neto Bs",
            "N.º transacción",
            "Banco",
            "Fecha/hora pago",
            "Estado del pago",
            "Comprobante",
            "Registrado por",
        ]
    ]

    st.dataframe(
        tabla,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Monto": st.column_config.NumberColumn(format="Bs %.2f"),
            "Comisión Bs": st.column_config.NumberColumn(format="Bs %.2f"),
            "Ingreso neto Bs": st.column_config.NumberColumn(format="Bs %.2f"),
        },
    )

    st.write("#### Revisar comprobante")

    opciones_comprobante = {
        (
            f"ID {int(fila.id)} · {fila.cliente} · "
            f"{fila.numero_transaccion or 'Sin transacción'}"
        ): int(fila.id)
        for fila in filtrado.itertuples()
        if getattr(fila, "comprobante", None) is not None
    }

    if opciones_comprobante:
        seleccion_comprobante = st.selectbox(
            "Selecciona una venta para revisar su comprobante",
            list(opciones_comprobante.keys()),
        )
        id_comprobante = opciones_comprobante[seleccion_comprobante]
        venta_comprobante = filtrado[
            filtrado["id"] == id_comprobante
        ].iloc[0]

        datos_archivo = venta_comprobante["comprobante"]
        mime_archivo = (
            venta_comprobante["comprobante_mime"]
            or "application/octet-stream"
        )
        nombre_archivo = (
            venta_comprobante["comprobante_nombre"]
            or f"comprobante_{id_comprobante}"
        )

        d1, d2 = st.columns([1.3, 0.7])
        with d1:
            if str(mime_archivo).startswith("image/"):
                st.image(
                    datos_archivo,
                    caption=(
                        f"Transacción: "
                        f"{venta_comprobante['numero_transaccion']}"
                    ),
                    use_container_width=True,
                )
            elif mime_archivo == "application/pdf":
                st.info(
                    "El comprobante es un PDF. Utiliza el botón de descarga "
                    "para abrirlo."
                )
            else:
                st.info("El archivo puede descargarse para su revisión.")

        with d2:
            st.write(
                f"**Banco:** {venta_comprobante['banco'] or '-'}"
            )
            st.write(
                f"**Estado:** "
                f"{venta_comprobante['estado_pago'] or 'Pendiente'}"
            )
            st.write(
                f"**Fecha/hora:** "
                f"{venta_comprobante['fecha_hora_pago'] or '-'}"
            )
            st.download_button(
                "Descargar comprobante",
                data=datos_archivo,
                file_name=nombre_archivo,
                mime=mime_archivo,
                use_container_width=True,
            )
    else:
        st.info("No hay comprobantes disponibles con los filtros actuales.")

# ------------------------------------------------------------
# EDITAR / ELIMINAR
# ------------------------------------------------------------

elif seccion == "Editar o eliminar":
    df = cargar_ventas()
    st.markdown('<div class="jet-section-title">Editar o eliminar una venta</div>', unsafe_allow_html=True)

    if df.empty:
        st.info("No existen ventas registradas.")
        st.stop()

    opciones = {
        f"ID {int(fila.id)} | {fila.fecha.date()} | {fila.curso} | "
        f"Bs {fila.monto:,.2f} | {fila.cliente}": int(fila.id)
        for fila in df.itertuples()
    }

    seleccion = st.selectbox("Selecciona una venta", list(opciones.keys()))
    id_venta = opciones[seleccion]
    registro = df[df["id"] == id_venta].iloc[0]

    with st.form("editar_venta"):
        c1, c2 = st.columns(2)

        with c1:
            fecha = st.date_input("Fecha", value=registro["fecha"].date())
            codigos_activos = list(COMISIONES.keys())
            curso_actual = str(registro["curso"])
            codigos_edicion = codigos_activos.copy()
            if curso_actual not in codigos_edicion:
                codigos_edicion.append(curso_actual)
            etiquetas_edicion = [
                etiqueta_curso(codigo) for codigo in codigos_edicion
            ]
            etiqueta_actual = etiqueta_curso(curso_actual)
            curso_etiqueta = st.selectbox(
                "Curso",
                etiquetas_edicion,
                index=etiquetas_edicion.index(etiqueta_actual),
            )
            curso = codigo_desde_etiqueta(curso_etiqueta)
            monto = st.number_input(
                "Monto",
                min_value=0.01,
                value=float(registro["monto"]),
                step=10.0,
            )

        with c2:
            cliente = st.text_input("Cliente", value=registro["cliente"])
            sede_actual = registro.get("sede", "") or "La Paz"
            opciones_sede = SEDES if sede_actual in SEDES else SEDES + [sede_actual]
            sede = st.selectbox(
                "Sede",
                opciones_sede,
                index=opciones_sede.index(sede_actual),
            )
            vendedor = st.text_input(
                "Nombre del vendedor",
                value=str(registro["vendedor"]),
                help="Puedes corregir o reemplazar libremente el nombre.",
            )
            tipo_cambio = st.number_input(
                "Tipo de cambio",
                min_value=0.01,
                value=float(registro["tipo_cambio"]),
                step=0.01,
            )

        actualizar = st.form_submit_button(
            "Actualizar venta",
            type="primary",
            use_container_width=True,
        )

        if actualizar:
            if not cliente.strip() or not vendedor.strip():
                st.error("Cliente y vendedor son obligatorios.")
            else:
                actualizar_venta(
                    id_venta,
                    fecha,
                    curso,
                    monto,
                    cliente,
                    vendedor,
                    sede,
                    tipo_cambio,
                )
                st.success("Venta actualizada correctamente.")

    st.divider()
    confirmar = st.checkbox("Confirmo que deseo eliminar esta venta")

    if st.button(
        "Eliminar venta",
        disabled=not confirmar,
        use_container_width=True,
    ):
        try:
            eliminar_venta(id_venta, usuario_actual["usuario"])
            st.success("Venta eliminada correctamente.")
            st.rerun()
        except ValueError as error:
            st.error(str(error))

# ------------------------------------------------------------
# LIMPIEZA DE DATOS
# ------------------------------------------------------------

elif seccion == "Limpieza de datos":
    st.markdown(
        '<div class="jet-section-title">Limpieza segura de datos</div>',
        unsafe_allow_html=True,
    )
    st.warning(
        "Esta herramienta es exclusiva para administradores. "
        "Puede eliminar ventas y registros de alumnos, pero conserva usuarios, "
        "cursos, bancos, sedes, tipo de cambio y configuración."
    )

    ventas_actuales, alumnos_actuales = obtener_conteos_limpieza()
    c1, c2 = st.columns(2)
    c1.metric("Ventas actuales", ventas_actuales)
    c2.metric("Alumnos actuales", alumnos_actuales)

    st.write("#### 1. Descarga un respaldo antes de limpiar")
    st.caption(
        "El respaldo contiene la base de datos completa, incluidos comprobantes, "
        "usuarios y configuraciones."
    )
    st.download_button(
        "Descargar respaldo de ventas.db",
        data=DB_PATH.read_bytes(),
        file_name=f"ventas_respaldo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
        mime="application/octet-stream",
        use_container_width=True,
    )

    st.divider()
    st.write("#### 2. Selecciona qué deseas eliminar")
    limpiar_ventas = st.checkbox(
        f"Eliminar todas las ventas ({ventas_actuales})",
        value=True,
    )
    limpiar_alumnos = st.checkbox(
        f"Eliminar todos los alumnos ({alumnos_actuales})",
        value=True,
    )

    if limpiar_ventas and not limpiar_alumnos and alumnos_actuales > 0:
        st.info(
            "Los alumnos se conservarán, pero quedarán como Con observación "
            "porque ya no tendrán una venta vinculada."
        )

    st.write("#### 3. Confirmación de seguridad")
    st.write(
        "Para habilitar el botón, escribe exactamente: "
        "**LIMPIAR DATOS**"
    )
    confirmacion_limpieza = st.text_input(
        "Frase de confirmación",
        placeholder="LIMPIAR DATOS",
    )
    confirmar_limpieza = st.checkbox(
        "Confirmo que descargué un respaldo y comprendo que esta acción no se puede deshacer."
    )

    habilitado = (
        confirmacion_limpieza.strip() == "LIMPIAR DATOS"
        and confirmar_limpieza
        and (limpiar_ventas or limpiar_alumnos)
    )

    if st.button(
        "Eliminar datos seleccionados",
        type="primary",
        disabled=not habilitado,
        use_container_width=True,
    ):
        try:
            ventas_borradas, alumnos_borrados = limpiar_datos_operativos(
                limpiar_ventas,
                limpiar_alumnos,
                usuario_actual["usuario"],
            )
            st.success(
                f"Limpieza completada: {ventas_borradas} venta(s) y "
                f"{alumnos_borrados} alumno(s) eliminados. "
                "La configuración general fue conservada."
            )
            st.rerun()
        except (ValueError, sqlite3.Error) as error:
            st.error(f"No se pudo completar la limpieza: {error}")

    st.divider()
    with st.expander("Historial reciente de acciones administrativas"):
        auditoria = cargar_auditoria()
        if auditoria.empty:
            st.info("Todavía no hay acciones registradas.")
        else:
            st.dataframe(auditoria, use_container_width=True, hide_index=True)

# ------------------------------------------------------------
# IMPORTAR
# ------------------------------------------------------------

elif seccion == "Importar Excel":
    st.markdown('<div class="jet-section-title">Importar ventas desde Excel</div>', unsafe_allow_html=True)
    st.write(
        "El archivo debe incluir: Fecha, Curso, Monto, Cliente y Vendedor. "
        "También puede incluir la columna Sede. "
        "También se acepta una columna llamada Tipo_ingreso."
    )

    tipo_cambio = st.number_input(
        "Tipo de cambio que se aplicará a las ventas importadas",
        min_value=0.01,
        value=float(TIPO_CAMBIO_USD_BS),
        step=0.01,
    )

    archivo = st.file_uploader(
        "Selecciona el Excel",
        type=["xlsx", "xlsm"],
    )

    if archivo is not None:
        vista = pd.read_excel(archivo, sheet_name="Ventas")
        st.write("Vista previa")
        st.dataframe(vista.head(20), use_container_width=True)

        if st.button("Importar ventas", type="primary"):
            try:
                archivo.seek(0)
                datos = pd.read_excel(archivo, sheet_name="Ventas")
                cantidad, errores = importar_dataframe(datos, tipo_cambio)
                st.success(f"Se importaron {cantidad} ventas correctamente.")

                if not errores.empty:
                    st.warning(
                        f"{len(errores)} filas no se importaron por contener errores."
                    )
                    st.dataframe(errores, use_container_width=True)
            except Exception as error:
                st.error(f"No se pudo importar el archivo: {error}")

# ------------------------------------------------------------
# EXPORTAR / RESPALDO
# ------------------------------------------------------------

elif seccion == "Exportar y respaldo":
    st.markdown('<div class="jet-section-title">Exportar reportes y respaldos</div>', unsafe_allow_html=True)
    df = cargar_ventas()

    if df.empty:
        st.info("No existen ventas para exportar.")
        st.stop()

    reporte = generar_reporte_excel(df)

    st.download_button(
        "Descargar reporte profesional en Excel",
        data=reporte,
        file_name="reporte_profesional_ventas_cursos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    csv = df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "Descargar respaldo CSV",
        data=csv,
        file_name="respaldo_ventas.csv",
        mime="text/csv",
        use_container_width=True,
    )

    with open(DB_PATH, "rb") as archivo_db:
        st.download_button(
            "Descargar copia completa de la base de datos",
            data=archivo_db.read(),
            file_name="ventas_respaldo.db",
            mime="application/octet-stream",
            use_container_width=True,
        )

# ------------------------------------------------------------
# CONFIGURACIÓN
# ------------------------------------------------------------

elif seccion == "Alumnos":
    st.markdown(
        '<div class="jet-section-title">Control de alumnos e inscripciones</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Compara los registros enviados por alumnos con las ventas del sistema. "
        "Enlace público: agrega ?registro=alumno al URL de esta aplicación."
    )

    alumnos_df = cargar_alumnos()
    ventas_df = cargar_ventas()

    total_ventas = len(ventas_df)
    total_alumnos = len(alumnos_df)
    validados = int((alumnos_df["estado"] == "Validado").sum()) if not alumnos_df.empty else 0
    observados = int((alumnos_df["estado"] == "Con observación").sum()) if not alumnos_df.empty else 0

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Ventas registradas", total_ventas)
    k2.metric("Alumnos registrados", total_alumnos)
    k3.metric("Validados", validados)
    k4.metric("Con observación", observados)

    enlace_base = st.text_input(
        "Enlace público para compartir",
        value="Agrega ?registro=alumno al final del enlace de Streamlit",
        disabled=True,
    )

    if alumnos_df.empty:
        st.info("Todavía no hay alumnos registrados mediante el formulario público.")
    else:
        f1, f2 = st.columns(2)
        with f1:
            filtro_estado = st.selectbox(
                "Estado",
                ["Todos", "Pendiente", "Validado", "Con observación", "Rechazado"],
            )
        with f2:
            buscar_alumno = st.text_input(
                "Buscar por nombre, documento o transacción"
            )

        vista = alumnos_df.copy()
        if filtro_estado != "Todos":
            vista = vista[vista["estado"] == filtro_estado]
        if buscar_alumno.strip():
            termino = buscar_alumno.strip()
            mascara = (
                vista["nombre_completo"].astype(str).str.contains(termino, case=False, na=False)
                | vista["documento"].astype(str).str.contains(termino, case=False, na=False)
                | vista["numero_transaccion"].astype(str).str.contains(termino, case=False, na=False)
            )
            vista = vista[mascara]

        tabla_alumnos = vista.rename(columns={
            "codigo": "Código",
            "nombre_completo": "Alumno",
            "documento": "Documento",
            "telefono": "Teléfono",
            "correo": "Correo",
            "numero_transaccion": "N.º transacción",
            "curso": "Curso",
            "sede": "Sede",
            "monto_reportado": "Monto",
            "estado": "Estado",
            "observacion": "Observación",
            "registrado_en": "Registrado",
            "vendedor": "Vendedor",
        })
        columnas = [
            "Código", "Alumno", "Documento", "Teléfono", "N.º transacción",
            "Curso", "Sede", "Monto", "Estado", "Vendedor", "Observación",
            "Registrado",
        ]
        st.dataframe(
            tabla_alumnos[columnas],
            use_container_width=True,
            hide_index=True,
            column_config={
                "Monto": st.column_config.NumberColumn(format="Bs %.2f"),
            },
        )

        st.write("### Revisar registro")
        opciones_alumnos = {
            f"{fila.codigo} · {fila.nombre_completo} · {fila.estado}": int(fila.id)
            for fila in alumnos_df.itertuples()
        }
        seleccionado = st.selectbox(
            "Selecciona un alumno",
            list(opciones_alumnos.keys()),
            key="alumno_revision",
        )
        id_alumno = opciones_alumnos[seleccionado]
        fila = alumnos_df[alumnos_df["id"] == id_alumno].iloc[0]

        r1, r2 = st.columns(2)
        with r1:
            st.write(f"**Alumno:** {fila['nombre_completo']}")
            st.write(f"**Documento:** {fila['documento']}")
            st.write(f"**Transacción:** {fila['numero_transaccion']}")
            st.write(f"**Curso/Sede:** {fila['curso'] or '-'} / {fila['sede'] or '-'}")
        with r2:
            st.write(f"**Estado actual:** {fila['estado']}")
            st.write(f"**Venta vinculada:** {'Sí' if pd.notna(fila['venta_id']) else 'No'}")
            st.write(f"**Vendedor:** {fila['vendedor'] if pd.notna(fila['vendedor']) else '-'}")
            st.write(f"**Observación:** {fila['observacion'] or '-'}")

        if pd.isna(fila["venta_id"]):
            if st.button("Buscar y vincular venta ahora", use_container_width=True):
                try:
                    intentar_vincular_alumno(id_alumno, usuario_actual["usuario"])
                    st.success("Venta vinculada y alumno validado.")
                    st.rerun()
                except ValueError as error:
                    st.error(str(error))

        with st.form("actualizar_estado_alumno_form"):
            estados = ["Pendiente", "Validado", "Con observación", "Rechazado"]
            estado_nuevo = st.selectbox(
                "Nuevo estado",
                estados,
                index=estados.index(fila["estado"]) if fila["estado"] in estados else 0,
            )
            observacion_nueva = st.text_area(
                "Observación administrativa",
                value=fila["observacion"] or "",
            )
            guardar_revision = st.form_submit_button(
                "Guardar revisión", type="primary", use_container_width=True
            )
            if guardar_revision:
                actualizar_estado_alumno(
                    id_alumno, estado_nuevo, observacion_nueva,
                    usuario_actual["usuario"],
                )
                st.success("Revisión guardada correctamente.")
                st.rerun()


elif seccion == "Usuarios":
    st.markdown('<div class="jet-section-title">Administración de usuarios</div>', unsafe_allow_html=True)

    tab_crear, tab_estado, tab_password = st.tabs(
        ["Crear usuario", "Activar o bloquear", "Cambiar contraseña"]
    )

    with tab_crear:
        with st.form("crear_usuario_form"):
            nombre_nuevo = st.text_input("Nombre completo")
            usuario_nuevo = st.text_input("Nombre de usuario")
            password_nuevo = st.text_input("Contraseña", type="password")
            rol_nuevo = st.selectbox("Rol", ROLES)
            crear_nuevo = st.form_submit_button(
                "Crear usuario",
                type="primary",
                use_container_width=True,
            )

            if crear_nuevo:
                try:
                    crear_usuario(
                        nombre_nuevo,
                        usuario_nuevo,
                        password_nuevo,
                        rol_nuevo,
                    )
                    st.success("Usuario creado correctamente.")
                except ValueError as error:
                    st.error(str(error))

    usuarios_df = cargar_usuarios()
    mapa_usuarios = {
        f"{fila.nombre} ({fila.usuario}) · {fila.rol}": int(fila.id)
        for fila in usuarios_df.itertuples()
    }

    with tab_estado:
        st.dataframe(
            usuarios_df.rename(
                columns={
                    "id": "ID",
                    "nombre": "Nombre",
                    "usuario": "Usuario",
                    "rol": "Rol",
                    "activo": "Activo",
                    "creado_en": "Creado",
                }
            ),
            use_container_width=True,
            hide_index=True,
        )

        seleccion_estado = st.selectbox(
            "Selecciona un usuario",
            list(mapa_usuarios.keys()),
            key="seleccion_estado_usuario",
        )
        id_estado = mapa_usuarios[seleccion_estado]
        fila_estado = usuarios_df[usuarios_df["id"] == id_estado].iloc[0]
        estado_actual = "Activo" if int(fila_estado["activo"]) == 1 else "Bloqueado"
        nuevo_estado = st.selectbox(
            "Estado",
            ["Activo", "Bloqueado"],
            index=0 if estado_actual == "Activo" else 1,
        )

        if st.button("Guardar estado", use_container_width=True):
            if id_estado == usuario_actual["id"] and nuevo_estado == "Bloqueado":
                st.error("No puedes bloquear tu propia cuenta mientras la usas.")
            else:
                cambiar_estado_usuario(id_estado, nuevo_estado == "Activo")
                st.success("Estado actualizado correctamente.")

    with tab_password:
        seleccion_password = st.selectbox(
            "Selecciona un usuario",
            list(mapa_usuarios.keys()),
            key="seleccion_password_usuario",
        )
        id_password = mapa_usuarios[seleccion_password]
        nueva_password = st.text_input("Nueva contraseña", type="password")
        confirmar_password = st.text_input(
            "Confirmar nueva contraseña",
            type="password",
        )

        if st.button(
            "Cambiar contraseña",
            type="primary",
            use_container_width=True,
        ):
            if nueva_password != confirmar_password:
                st.error("Las contraseñas no coinciden.")
            else:
                try:
                    cambiar_password_usuario(id_password, nueva_password)
                    st.success("Contraseña actualizada correctamente.")
                except ValueError as error:
                    st.error(str(error))


elif seccion == "Configuración":
    st.markdown(
        '<div class="jet-section-title">Configuración del sistema</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Los cambios realizados aquí se guardan en la base de datos y se "
        "aplican automáticamente en todo el sistema."
    )

    tab_cursos, tab_sedes, tab_bancos, tab_cambio = st.tabs(
        ["📚 Cursos", "🏢 Sedes", "🏦 Bancos", "💱 Tipo de cambio"]
    )

    # --------------------------------------------------------
    # CURSOS
    # --------------------------------------------------------
    with tab_cursos:
        st.write("### Cursos y comisiones")
        cursos_df = cargar_cursos(solo_activos=False)

        vista_cursos = cursos_df.copy()
        if not vista_cursos.empty:
            vista_cursos["Estado"] = vista_cursos["activo"].map(
                {1: "Activo", 0: "Inactivo"}
            )
            vista_cursos["Ventas"] = vista_cursos["codigo"].apply(
                cantidad_ventas_curso
            )
            vista_cursos = vista_cursos.rename(
                columns={
                    "codigo": "Código",
                    "nombre": "Nombre",
                    "comision": "Comisión",
                    "moneda": "Moneda",
                }
            )[["Código", "Nombre", "Comisión", "Moneda", "Estado", "Ventas"]]
            st.dataframe(
                vista_cursos,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Comisión": st.column_config.NumberColumn(format="%.2f"),
                },
            )

        sub_nuevo, sub_editar = st.tabs(
            ["➕ Crear curso", "✏️ Editar, desactivar o eliminar"]
        )

        with sub_nuevo:
            with st.form("crear_curso_form", clear_on_submit=True):
                n1, n2 = st.columns(2)
                with n1:
                    nuevo_codigo = st.text_input(
                        "Código del curso",
                        placeholder="Ej.: PCA",
                    )
                    nueva_comision = st.number_input(
                        "Comisión",
                        min_value=0.0,
                        value=0.0,
                        step=10.0,
                    )
                with n2:
                    nuevo_nombre = st.text_input(
                        "Nombre del curso",
                        placeholder="Ej.: Piloto Comercial Avanzado",
                    )
                    nueva_moneda = st.selectbox("Moneda", ["USD", "BOB"])

                nuevo_activo = st.checkbox("Curso activo", value=True)
                crear_curso_btn = st.form_submit_button(
                    "Crear curso",
                    type="primary",
                    use_container_width=True,
                )

                if crear_curso_btn:
                    try:
                        crear_curso(
                            nuevo_codigo,
                            nuevo_nombre,
                            nueva_comision,
                            nueva_moneda,
                            nuevo_activo,
                        )
                        st.success("Curso creado correctamente.")
                        st.rerun()
                    except ValueError as error:
                        st.error(str(error))

        with sub_editar:
            cursos_df = cargar_cursos(solo_activos=False)
            if cursos_df.empty:
                st.info("No hay cursos para editar.")
            else:
                opciones = {
                    f"{fila.codigo} · {fila.nombre}": fila.codigo
                    for fila in cursos_df.itertuples()
                }
                seleccion_curso = st.selectbox(
                    "Selecciona un curso",
                    list(opciones.keys()),
                    key="curso_config_seleccion",
                )
                codigo_actual = opciones[seleccion_curso]
                fila = cursos_df[
                    cursos_df["codigo"] == codigo_actual
                ].iloc[0]

                with st.form("editar_curso_form"):
                    e1, e2 = st.columns(2)
                    with e1:
                        codigo_editado = st.text_input(
                            "Código",
                            value=str(fila["codigo"]),
                        )
                        comision_editada = st.number_input(
                            "Comisión",
                            min_value=0.0,
                            value=float(fila["comision"]),
                            step=10.0,
                        )
                    with e2:
                        nombre_editado = st.text_input(
                            "Nombre",
                            value=str(fila["nombre"]),
                        )
                        moneda_editada = st.selectbox(
                            "Moneda",
                            ["USD", "BOB"],
                            index=0 if fila["moneda"] == "USD" else 1,
                        )

                    activo_editado = st.checkbox(
                        "Curso activo",
                        value=bool(fila["activo"]),
                    )
                    guardar_curso_btn = st.form_submit_button(
                        "Guardar cambios",
                        type="primary",
                        use_container_width=True,
                    )

                    if guardar_curso_btn:
                        try:
                            actualizar_curso(
                                codigo_actual,
                                codigo_editado,
                                nombre_editado,
                                comision_editada,
                                moneda_editada,
                                activo_editado,
                            )
                            st.success("Curso actualizado correctamente.")
                            st.rerun()
                        except ValueError as error:
                            st.error(str(error))

                ventas_asociadas = cantidad_ventas_curso(codigo_actual)
                if ventas_asociadas:
                    st.info(
                        f"Este curso tiene {ventas_asociadas} venta(s). "
                        "No puede eliminarse, pero sí desactivarse."
                    )
                confirmar_eliminar_curso = st.checkbox(
                    "Confirmo que deseo eliminar este curso",
                    key="confirmar_eliminar_curso",
                    disabled=ventas_asociadas > 0,
                )
                if st.button(
                    "Eliminar curso",
                    disabled=(
                        ventas_asociadas > 0
                        or not confirmar_eliminar_curso
                    ),
                    use_container_width=True,
                ):
                    try:
                        eliminar_curso(codigo_actual)
                        st.success("Curso eliminado correctamente.")
                        st.rerun()
                    except ValueError as error:
                        st.error(str(error))

    # --------------------------------------------------------
    # CATÁLOGOS: SEDES Y BANCOS
    # --------------------------------------------------------
    def mostrar_gestion_catalogo(tabla, singular):
        df_catalogo = cargar_catalogo(tabla, solo_activos=False)
        vista = df_catalogo.copy()
        if not vista.empty:
            vista["Estado"] = vista["activo"].map(
                {1: "Activo", 0: "Inactivo"}
            )
            vista = vista.rename(
                columns={"id": "ID", "nombre": "Nombre"}
            )[["ID", "Nombre", "Estado"]]
            st.dataframe(vista, use_container_width=True, hide_index=True)

        with st.form(f"crear_{tabla}_form", clear_on_submit=True):
            nombre_nuevo = st.text_input(
                f"Nombre de la nueva {singular.lower()}",
                key=f"nuevo_{tabla}",
            )
            crear_btn = st.form_submit_button(
                f"Agregar {singular.lower()}",
                type="primary",
                use_container_width=True,
            )
            if crear_btn:
                try:
                    crear_elemento_catalogo(tabla, nombre_nuevo)
                    st.success(f"{singular} agregada correctamente.")
                    st.rerun()
                except ValueError as error:
                    st.error(str(error))

        df_catalogo = cargar_catalogo(tabla, solo_activos=False)
        if not df_catalogo.empty:
            mapa = {
                f"{fila.nombre} · {'Activo' if fila.activo else 'Inactivo'}":
                int(fila.id)
                for fila in df_catalogo.itertuples()
            }
            seleccionado = st.selectbox(
                f"Selecciona una {singular.lower()} para editar",
                list(mapa.keys()),
                key=f"seleccion_{tabla}",
            )
            id_elemento = mapa[seleccionado]
            fila = df_catalogo[df_catalogo["id"] == id_elemento].iloc[0]

            with st.form(f"editar_{tabla}_form"):
                nombre_editado = st.text_input(
                    "Nombre",
                    value=str(fila["nombre"]),
                    key=f"nombre_editar_{tabla}",
                )
                activo_editado = st.checkbox(
                    "Activo",
                    value=bool(fila["activo"]),
                    key=f"activo_editar_{tabla}",
                )
                guardar_btn = st.form_submit_button(
                    "Guardar cambios",
                    type="primary",
                    use_container_width=True,
                )
                if guardar_btn:
                    try:
                        actualizar_elemento_catalogo(
                            tabla,
                            id_elemento,
                            nombre_editado,
                            activo_editado,
                        )
                        st.success("Cambios guardados correctamente.")
                        st.rerun()
                    except ValueError as error:
                        st.error(str(error))

            confirmar = st.checkbox(
                f"Confirmo que deseo eliminar esta {singular.lower()}",
                key=f"confirmar_eliminar_{tabla}",
            )
            if st.button(
                f"Eliminar {singular.lower()}",
                disabled=not confirmar,
                key=f"eliminar_{tabla}",
                use_container_width=True,
            ):
                try:
                    eliminar_elemento_catalogo(tabla, id_elemento)
                    st.success(f"{singular} eliminada correctamente.")
                    st.rerun()
                except ValueError as error:
                    st.error(str(error))

    with tab_sedes:
        st.write("### Gestión de sedes")
        st.caption(
            "Las sedes inactivas dejan de aparecer en nuevas ventas, "
            "pero permanecen en el historial."
        )
        mostrar_gestion_catalogo("sedes", "Sede")

    with tab_bancos:
        st.write("### Gestión de bancos y medios de pago")
        st.caption(
            "Los elementos inactivos dejan de aparecer en el formulario "
            "de nuevas ventas."
        )
        mostrar_gestion_catalogo("bancos", "Opción")

    # --------------------------------------------------------
    # TIPO DE CAMBIO
    # --------------------------------------------------------
    with tab_cambio:
        st.write("### Tipo de cambio predeterminado")
        st.caption(
            "Este valor aparecerá automáticamente al registrar o importar "
            "nuevas ventas. Cada venta conserva el tipo de cambio utilizado."
        )

        with st.form("tipo_cambio_form"):
            nuevo_tipo_cambio = st.number_input(
                "USD/BOB",
                min_value=0.01,
                value=float(obtener_tipo_cambio()),
                step=0.01,
                format="%.2f",
            )
            guardar_cambio_btn = st.form_submit_button(
                "Actualizar tipo de cambio",
                type="primary",
                use_container_width=True,
            )

            if guardar_cambio_btn:
                try:
                    guardar_tipo_cambio(nuevo_tipo_cambio)
                    st.success("Tipo de cambio actualizado correctamente.")
                    st.rerun()
                except ValueError as error:
                    st.error(str(error))

        st.warning(
            "Los cambios de comisión y tipo de cambio se aplican a ventas "
            "nuevas. Las ventas anteriores conservan sus valores históricos."
        )

    st.warning(
        "El acceso inicial es admin / Admin123!. "
        "Cambia esa contraseña desde Usuarios."
    )
